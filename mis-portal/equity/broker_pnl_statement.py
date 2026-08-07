"""Broker realised-P&L statement parsers — Zerodha Console, Angel One, Dhan.

Each broker publishes an FY-aggregated realised-P&L statement: one row per scrip
with Buy Value / Sell Value / Realised P&L for the window. These are the broker's
own authority for capital gains (computed off the depository's true corporate-action
history), so we ingest them as a per-scrip ORACLE to reconcile our FIFO engine
against — NOT as a source of trades (there are no dated fills here).

Three wire formats, one normalised shape. `detect()` sniffs the broker cheaply for
the upload guard; `parse()` opens the file, auto-detects, and dispatches. The
per-scrip `realised_pnl` we keep is always the GROSS price-to-price figure, because
that is what report_generator._fifo_realised_grouped produces (charges are listed
separately and not deducted there).

parse() returns:
    {broker, client_id, period_from (date), period_to (date), fy_label,
     downloaded_at (datetime|None), segment_totals {SEG: {realised, net, charges}},
     lines [ {segment, security_name, isin, quantity, buy_value, sell_value,
              realised_pnl, st_pnl, lt_pnl, return_pct} ] }
"""
from __future__ import annotations

import csv
import re
from datetime import datetime, date
from decimal import Decimal, InvalidOperation

import openpyxl


# ---------------------------------------------------------------- small helpers
def _num(v):
    """'1,200' / '13,068.00' / '"0.00"' / None → float | None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace('"', "").strip()
    if s in ("", "-", "nan", "None", "N/A"):
        return None
    try:
        return float(Decimal(s))
    except (InvalidOperation, ValueError):
        return None


def _date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v or "").strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %b %Y", "%d-%b-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _fy_label(d: date | None) -> str | None:
    """Indian FY (Apr→Mar) label for the period start, e.g. 2024-06-01 → 'FY24-25'.
    Returned only when period_from lands cleanly on an FY start month bucket; a
    statement is still tagged by the FY its start date falls in."""
    if not d:
        return None
    y = d.year if d.month >= 4 else d.year - 1
    return f"FY{str(y)[2:]}-{str(y + 1)[2:]}"


def _seg_of_name(name: str) -> str:
    """Dhan/Angel carry derivatives inline; classify by the scrip-name prefix.
    'FUT NATURALGAS ...' / 'OPT PETRONET ... PE' → FnO, everything else EQ."""
    n = (name or "").strip().upper()
    return "FnO" if n.startswith(("FUT ", "OPT ", "FUT-", "OPT-")) else "EQ"


# ------------------------------------------------------------------- detection
def detect(filename: str, head: str = "") -> str | None:
    """Best-effort broker sniff for the upload guard. Content wins over filename;
    xlsx (binary head) falls back to filename hints — parse() re-verifies by
    structure regardless, so a wrong guess here only affects the early reject."""
    h = (head or "").lower()
    if "realised pnl report" in h or ("scrip name" in h and "realised p&l" in h):
        return "dhan"
    fn = (filename or "").lower()
    if fn.endswith(".csv") and "dhan" in fn:
        return "dhan"
    if fn.endswith(".xlsx"):
        if "profitloss_statement" in fn:
            return "angel_one"
        if fn.startswith("pnl-") or "pnl-" in fn:
            return "zerodha"
    return None


# --------------------------------------------------------------------- parsers
def parse(path: str) -> dict:
    """Open `path`, auto-detect the broker by structure, and parse."""
    lower = path.lower()
    if lower.endswith(".csv"):
        with open(path, "r", encoding="utf-8-sig", errors="ignore") as fh:
            text = fh.read()
        if "realised pnl report" in text.lower():
            return _parse_dhan(text, path)
        raise ValueError("Unrecognised CSV — not a Dhan realised-P&L report.")

    wb = openpyxl.load_workbook(path, data_only=True)
    names = set(wb.sheetnames)
    if "Equity P&L" in names or "F&O P&L" in names:
        return _parse_angel(wb, path)
    if "Equity" in names:
        # Zerodha Console: the 'Equity' sheet carries "P&L Statement for Equity from".
        ws = wb["Equity"]
        for row in ws.iter_rows(min_row=1, max_row=14, values_only=True):
            if any(isinstance(c, str) and "P&L Statement" in c for c in row):
                return _parse_zerodha(wb, path)
    raise ValueError(f"Unrecognised statement layout (sheets: {sorted(names)}).")


def parse_bytes(filename: str, data: bytes) -> dict:
    """parse() for an in-memory upload — spools to a temp file (openpyxl/csv both
    want a path) and cleans up."""
    import tempfile, os
    suffix = ".csv" if (filename or "").lower().endswith(".csv") else ".xlsx"
    with tempfile.NamedTemporaryFile("wb", suffix=suffix, delete=False) as tf:
        tf.write(data)
        tmp = tf.name
    try:
        return parse(tmp)
    finally:
        os.unlink(tmp)


def _parse_zerodha(wb, path) -> dict:
    """Zerodha Console 'pnl-<ID>.xlsx'. Client ID + period from the preamble; the
    per-scrip table (Symbol|ISIN|Quantity|Buy Value|Sell Value|Realized P&L|…) is
    found by scanning for its header. Realized P&L is gross (charges are a separate
    block). Only rows that actually realised (Quantity>0 or P&L≠0) are kept."""
    ws = wb["Equity"]
    rows = list(ws.iter_rows(values_only=True))

    client_id = None
    period_from = period_to = None
    charges = None
    realised_total = None
    for r in rows[:20]:
        cells = [c for c in r if c is not None]
        for i, c in enumerate(cells):
            s = str(c).strip()
            if s == "Client ID" and i + 1 < len(cells):
                client_id = str(cells[i + 1]).strip()
            elif s.startswith("P&L Statement") and "from" in s:
                m = re.search(r"from\s+(\d{4}-\d{2}-\d{2})\s+to\s+(\d{4}-\d{2}-\d{2})", s)
                if m:
                    period_from, period_to = _date(m.group(1)), _date(m.group(2))
            elif s == "Charges" and i + 1 < len(cells):
                charges = _num(cells[i + 1])
            elif s == "Realized P&L" and i + 1 < len(cells):
                realised_total = _num(cells[i + 1])

    # Locate the per-scrip header row.
    hdr_idx = None
    for i, r in enumerate(rows):
        vals = [str(c).strip() if c is not None else "" for c in r]
        if "Symbol" in vals and "ISIN" in vals and "Realized P&L" in vals:
            hdr_idx, hdr = i, vals
            break
    if hdr_idx is None:
        raise ValueError("Zerodha statement: per-scrip Symbol/ISIN header not found.")
    col = {name: j for j, name in enumerate(hdr) if name}

    def cell(r, name):
        j = col.get(name)
        return r[j] if j is not None and j < len(r) else None

    lines = []
    for r in rows[hdr_idx + 1:]:
        sym = cell(r, "Symbol")
        if sym is None or not str(sym).strip():
            continue
        qty = _num(cell(r, "Quantity")) or 0.0
        rp = _num(cell(r, "Realized P&L"))
        if qty == 0 and (rp is None or rp == 0):
            continue  # pure open holding — no realisation this window
        lines.append({
            "segment":      "EQ",
            "security_name": str(sym).strip(),
            "isin":         (str(cell(r, "ISIN")).strip() or None) if cell(r, "ISIN") else None,
            "quantity":     qty,
            "buy_value":    _num(cell(r, "Buy Value")),
            "sell_value":   _num(cell(r, "Sell Value")),
            "realised_pnl": rp or 0.0,
            "st_pnl":       None,
            "lt_pnl":       None,
            "return_pct":   _num(cell(r, "Realized P&L Pct.")),
        })

    eq_realised = realised_total if realised_total is not None else sum(l["realised_pnl"] for l in lines)
    return {
        "broker": "zerodha", "client_id": client_id,
        "period_from": period_from, "period_to": period_to,
        "fy_label": _fy_label(period_from), "downloaded_at": None,
        "segment_totals": {"EQ": {"realised": round(eq_realised, 2), "charges": charges}},
        "lines": lines,
    }


def _parse_angel(wb, path) -> dict:
    """Angel One 'ProfitLoss_Statement_<ID>.xlsx'. Two sheets: 'Equity P&L'
    (a Delivery per-scrip table with Gross/Net + ST/LT) and 'F&O P&L'. We keep the
    delivery rows' GROSS PnL (to match our gross FIFO) and record intraday/net from
    the summary. F&O is stored as segment FnO."""
    eq_ws = wb["Equity P&L"] if "Equity P&L" in wb.sheetnames else None
    fno_ws = wb["F&O P&L"] if "F&O P&L" in wb.sheetnames else None

    client_id = period_from = period_to = downloaded_at = None
    seg_totals = {}
    lines = []

    def read_preamble(ws):
        nonlocal client_id, period_from, period_to, downloaded_at
        for r in ws.iter_rows(min_row=1, max_row=13, values_only=True):
            cells = [c for c in r if c is not None]
            for i, c in enumerate(cells):
                s = str(c).strip()
                if s in ("Client Id", "Client ID") and i + 1 < len(cells):
                    client_id = str(cells[i + 1]).strip()
                elif s == "From Date" and i + 1 < len(cells):
                    period_from = _date(cells[i + 1])
                elif s == "To Date" and i + 1 < len(cells):
                    period_to = _date(cells[i + 1])
                elif s == "Date Of Download" and i + 1 < len(cells):
                    downloaded_at = datetime.combine(_date(cells[i + 1]) or date.today(), datetime.min.time())

    def parse_scrip_table(ws, segment):
        rows = list(ws.iter_rows(values_only=True))
        hdr_idx = None
        for i, r in enumerate(rows):
            vals = [str(c).strip() if c is not None else "" for c in r]
            if ("Scrip Symbol" in vals) and ("Gross PnL" in vals or "Net PnL" in vals):
                hdr_idx, hdr = i, vals
                break
        if hdr_idx is None:
            return
        col = {name: j for j, name in enumerate(hdr) if name}

        def cell(r, name):
            j = col.get(name)
            return r[j] if j is not None and j < len(r) else None

        for r in rows[hdr_idx + 1:]:
            sym = cell(r, "Scrip Symbol")
            if sym is None or not str(sym).strip():
                continue
            if str(sym).strip().lower() in ("total", "grand total"):
                continue
            gross = _num(cell(r, "Gross PnL"))
            buy_v = _num(cell(r, "Buy Value"))
            sell_v = _num(cell(r, "Sell Value"))
            # Rights-entitlement / bonus placeholder rows come through as all-zero.
            if (gross in (None, 0)) and (buy_v in (None, 0)) and (sell_v in (None, 0)):
                continue
            name = cell(r, "Company Name") or sym
            lines.append({
                "segment":       segment,
                "security_name": str(name).strip(),
                "isin":          None,   # Angel gives no ISIN
                "quantity":      _num(cell(r, "Quantity")),
                "buy_value":     buy_v,
                "sell_value":    sell_v,
                "realised_pnl":  gross or 0.0,
                "st_pnl":        _num(cell(r, "Short term PnL")),
                "lt_pnl":        _num(cell(r, "Long term PnL")),
                "return_pct":    None,
            })

    if eq_ws is not None:
        read_preamble(eq_ws)
        parse_scrip_table(eq_ws, "EQ")
        # Summary figures for context (net incl. intraday; gross of the delivery rows).
        summ = _angel_summary(eq_ws)
        seg_totals["EQ"] = {
            "realised": round(sum(l["realised_pnl"] for l in lines if l["segment"] == "EQ"), 2),
            "net":      summ.get("Net PnL"),
            "charges":  summ.get("Total Brokerage"),
            "intraday_net": summ.get("Intraday Net PnL"),
        }
    if fno_ws is not None:
        if client_id is None:
            read_preamble(fno_ws)
        before = len(lines)
        parse_scrip_table(fno_ws, "FnO")
        if len(lines) > before:
            summ = _angel_summary(fno_ws)
            seg_totals["FnO"] = {
                "realised": round(sum(l["realised_pnl"] for l in lines if l["segment"] == "FnO"), 2),
                "net":      summ.get("Net PnL"),
            }

    return {
        "broker": "angel_one", "client_id": client_id,
        "period_from": period_from, "period_to": period_to,
        "fy_label": _fy_label(period_from), "downloaded_at": downloaded_at,
        "segment_totals": seg_totals, "lines": lines,
    }


def _angel_summary(ws) -> dict:
    """The Angel 'X | value' summary lines above the per-scrip table (Net PnL,
    Total Brokerage, Intraday Net PnL, …)."""
    out = {}
    for r in ws.iter_rows(min_row=1, max_row=30, values_only=True):
        cells = [c for c in r if c is not None]
        if len(cells) >= 2 and isinstance(cells[0], str):
            out[cells[0].strip()] = _num(cells[1])
    return out


def _parse_dhan(text: str, path) -> dict:
    """Dhan 'Realised P&L' CSV: metadata preamble, a blank line, then the per-scrip
    table (Scrip Name, …, Realised P&L, %), then a Net/Gross footer. Per-scrip
    Realised P&L is gross; FUT/OPT names are FnO."""
    rows = list(csv.reader(text.splitlines()))
    client_id = period_from = period_to = downloaded_at = None
    gross_total = net_total = charges = None

    for r in rows[:8]:
        if not r:
            continue
        key = (r[0] or "").strip().lower()
        if key == "realised pnl report" and len(r) > 1:
            m = re.search(r"from\s+(\d{2}-\d{2}-\d{4})\s+to\s+(\d{2}-\d{2}-\d{4})", r[1], re.I)
            if m:
                period_from, period_to = _date(m.group(1)), _date(m.group(2))
        elif key == "ucc" and len(r) > 1:
            client_id = (r[1] or "").strip()

    # Locate the per-scrip header + the footer.
    hdr_idx = None
    for i, r in enumerate(rows):
        if r and (r[0] or "").strip().lower() == "scrip name":
            hdr_idx, hdr = i, [c.strip() for c in r]
            break
    if hdr_idx is None:
        raise ValueError("Dhan statement: 'Scrip Name' header not found.")
    col = {name: j for j, name in enumerate(hdr)}

    def cell(r, name):
        j = col.get(name)
        return r[j] if j is not None and j < len(r) else None

    lines = []
    for r in rows[hdr_idx + 1:]:
        if not r or not (r[0] or "").strip():
            continue
        name = (r[0] or "").strip()
        low = name.lower()
        if low.startswith("net p&l"):
            # footer: Net P&L,<v>,Brokerage,<v>,Gross P&L,<v>,Total Charges,<v>
            kv = {(r[k] or "").strip().lower(): _num(r[k + 1]) for k in range(0, len(r) - 1, 2)}
            net_total = kv.get("net p&l"); gross_total = kv.get("gross p&l")
            charges = kv.get("total charges")
            continue
        if low.startswith("note"):
            m = re.search(r"downloaded at\s+(.+)$", name, re.I)
            if m:
                try:
                    downloaded_at = datetime.strptime(m.group(1).strip(), "%m/%d/%Y %I:%M %p")
                except ValueError:
                    downloaded_at = None
            continue
        rp = _num(cell(r, "Realised P&L"))
        if rp is None:
            continue
        lines.append({
            "segment":       _seg_of_name(name),
            "security_name": name,
            "isin":          None,
            "quantity":      _num(cell(r, "Quantity")),
            "buy_value":     _num(cell(r, "Buy Value")),
            "sell_value":    _num(cell(r, "Sell Value")),
            "realised_pnl":  rp,
            "st_pnl":        None,
            "lt_pnl":        None,
            "return_pct":    _num(cell(r, "Realised P&L %")),
        })

    seg_totals = {}
    for seg in ("EQ", "FnO"):
        s = [l for l in lines if l["segment"] == seg]
        if s:
            seg_totals[seg] = {"realised": round(sum(l["realised_pnl"] for l in s), 2)}
    if gross_total is not None:
        seg_totals.setdefault("EQ", {})["gross_all"] = gross_total
        seg_totals.setdefault("EQ", {})["net_all"] = net_total
        seg_totals.setdefault("EQ", {})["charges"] = charges

    return {
        "broker": "dhan", "client_id": client_id,
        "period_from": period_from, "period_to": period_to,
        "fy_label": _fy_label(period_from), "downloaded_at": downloaded_at,
        "segment_totals": seg_totals, "lines": lines,
    }


if __name__ == "__main__":
    import sys, json
    res = parse(sys.argv[1])
    print(f"broker={res['broker']} client={res['client_id']} "
          f"{res['period_from']}→{res['period_to']} fy={res['fy_label']} "
          f"downloaded={res['downloaded_at']}")
    print("segment_totals:", json.dumps(res["segment_totals"]))
    print(f"lines={len(res['lines'])}  "
          f"EQ={sum(1 for l in res['lines'] if l['segment']=='EQ')}  "
          f"FnO={sum(1 for l in res['lines'] if l['segment']=='FnO')}")
    tot = sum(l["realised_pnl"] for l in res["lines"])
    print(f"sum(line realised_pnl) = {tot:,.2f}")
    for l in res["lines"][:6]:
        print(f"  [{l['segment']}] {l['security_name'][:26]:26} isin={l['isin']} "
              f"qty={l['quantity']} rpnl={l['realised_pnl']:,.2f}")
