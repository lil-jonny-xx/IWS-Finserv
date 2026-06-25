#!/usr/bin/env python3
"""
Broker cash-ledger importer.

Parses Zerodha / Dhan / Angel One cash ledgers (and Vested's dividend Income sheet)
and populates:
  • cash_ledger        — daily closing cash balance per (entity, broker-account).
  • external_cashflow  — classified non-trade cash: DEPOSIT / WITHDRAWAL (for portfolio
                         money-weighted return) and DIVIDEND / INTEREST (to enrich
                         holding XIRR). Trade settlements & charges are intentionally
                         skipped (trades are in stock_transaction; charges are XIRR noise).

Sign convention for external_cashflow.amount_native (investor's perspective):
  deposit into broker = -outflow ; withdrawal to bank / dividend = +inflow.

Auto-detects broker + entity from filename, applying the known Dhan mislabel fix
(client ***REMOVED*** = HHR, not "Rajani Corp"). Dry-run by default; --commit to write.
"""
import os
import sys
import csv
import glob
import argparse
import hashlib
import warnings
from datetime import datetime

import openpyxl
import psycopg2
from psycopg2.extras import RealDictCursor
from dotenv import load_dotenv

warnings.filterwarnings("ignore")
load_dotenv("/var/www/mis-portal/.env", override=True)

INCOMING = "/var/www/mis-portal/data/incoming/tradebooks_ledgers/Tradebooks and Ledgers"

# External (bank↔broker) cash only. NOT inter-segment transfers like
# "BEING AMT TRF FROM NSE / TO MCX", which are internal and must be excluded.
DEPOSIT_HINTS = ("funds added", "funds deposited", "money added")
WITHDRAW_HINTS = ("funds transferred back", "payout", "pay out", "auto-settled",
                  "quarterly settlement", "withdrawal")


def _num(v):
    if v is None:
        return 0.0
    s = str(v).replace(",", "").strip().strip('"')
    if s in ("", "-", "None", "nan"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _date(v):
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip().strip('"')
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%d-%b-%Y"):
        try:
            return datetime.strptime(s[:len(fmt) + 4], fmt).date()
        except ValueError:
            continue
    return None


def classify(text):
    t = (text or "").lower()
    if any(h in t for h in DEPOSIT_HINTS):
        return "DEPOSIT"
    if any(h in t for h in WITHDRAW_HINTS):
        return "WITHDRAWAL"
    if "dividend" in t:
        return "DIVIDEND"
    if "interest" in t:
        return "INTEREST"
    return None


# ---- parsers: yield normalised entries {date, text, debit, credit, balance} ----
def _excel_rows(path, sheet=0):
    """All cells of an Excel sheet as a list of row-lists (handles .xls via xlrd and
    .xlsx via openpyxl, both through pandas)."""
    import pandas as pd
    df = pd.read_excel(path, sheet_name=sheet, header=None, dtype=object)
    return df.where(df.notna(), None).values.tolist()


def _hdr_index(rows, *required):
    for i, r in enumerate(rows):
        cells = {str(c).strip() for c in r if c is not None}
        if all(req in cells for req in required):
            return i, {str(c).strip(): j for j, c in enumerate(r) if c is not None}
    return None, None


def parse_zerodha_ledger(path):
    # Zerodha ships the ledger as CSV or as an .xlsx with a preamble (header row
    # "Particulars | Posting Date | … | Net Balance"). Same columns either way.
    if path.lower().endswith((".xlsx", ".xls")):
        rows = _excel_rows(path)
        hdr, col = _hdr_index(rows, "Particulars", "Posting Date")
        if hdr is None:
            return
        for r in rows[hdr + 1:]:
            d = _date(r[col["Posting Date"]]) if "Posting Date" in col else None
            if not d:
                continue
            yield {"date": d, "text": str(r[col["Particulars"]] or ""),
                   "debit": _num(r[col.get("Debit")]) if "Debit" in col else 0.0,
                   "credit": _num(r[col.get("Credit")]) if "Credit" in col else 0.0,
                   "balance": _num(r[col.get("Net Balance")]) if "Net Balance" in col else 0.0}
        return
    with open(path, newline="") as fh:
        for r in csv.DictReader(fh):
            d = _date(r.get("posting_date"))
            if not d:
                continue
            yield {"date": d, "text": r.get("particulars", ""),
                   "debit": _num(r.get("debit")), "credit": _num(r.get("credit")),
                   "balance": _num(r.get("net_balance"))}


def parse_dhan_ledger(path):
    # Dhan ships two layouts: the older CSV ("Posting Date" header) and the "Ledger
    # Statement" .xls (preamble + "Sr|Date|Transaction ID|Type|Description|…|Running
    # Balance"). Classification text = Type + Description so "Funds Deposited" matches.
    if path.lower().endswith((".xls", ".xlsx")):
        rows = _excel_rows(path)
        hdr, col = _hdr_index(rows, "Date", "Running Balance")
        if hdr is None:
            return
        for r in rows[hdr + 1:]:
            d = _date(r[col["Date"]]) if "Date" in col else None
            if not d:
                continue
            typ = str(r[col["Type"]] or "") if "Type" in col else ""
            desc = str(r[col["Description"]] or "") if "Description" in col else ""
            yield {"date": d, "text": f"{typ} {desc}".strip(),
                   "debit": _num(r[col.get("Debit")]) if "Debit" in col else 0.0,
                   "credit": _num(r[col.get("Credit")]) if "Credit" in col else 0.0,
                   "balance": _num(r[col.get("Running Balance")]) if "Running Balance" in col else 0.0}
        return
    with open(path, newline="") as fh:
        rows = list(csv.reader(fh))
    hdr = next((i for i, r in enumerate(rows) if r and r[0].strip() == "Posting Date"), None)
    if hdr is None:
        return
    for r in rows[hdr + 1:]:
        if len(r) < 7:
            continue
        d = _date(r[0])
        if not d:
            continue
        yield {"date": d, "text": r[2], "credit": _num(r[4]), "debit": _num(r[5]),
               "balance": _num(r[6])}


def parse_angel_ledger(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Broking Ledger"]
    rows = [tuple(x) for x in ws.iter_rows(values_only=True)]
    wb.close()
    hdr = next((i for i, r in enumerate(rows)
                if r and str(r[0]).strip() == "Transaction"), None)
    if hdr is None:
        return
    for r in rows[hdr + 1:]:
        if not r or not r[0]:
            continue
        d = _date(r[1])
        if not d:
            continue
        yield {"date": d, "text": str(r[0]), "debit": _num(r[4]), "credit": _num(r[5]),
               "balance": _num(r[6])}


def parse_vested_income(path):
    """Vested dividends/interest (USD) from the Income sheet → DIVIDEND/INTEREST rows."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Income"]
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        d = _date(row[0]); activity = str(row[2] or "").strip(); ticker = row[3]; amt = _num(row[4])
        if not d or amt == 0:
            continue
        ftype = "DIVIDEND" if "dividend" in activity.lower() else \
                "INTEREST" if "interest" in activity.lower() else \
                "DIVIDEND" if activity.lower() == "tax" else None  # Tax is the div-withholding leg
        if ftype:
            out.append({"date": d, "symbol": str(ticker) if ticker else None,
                        "amount": amt, "activity": activity, "flow_type": ftype})
    wb.close()
    return out


LEDGER_PARSERS = {"zerodha": parse_zerodha_ledger, "dhan": parse_dhan_ledger,
                  "angel_one": parse_angel_ledger}


def detect(path):
    """(broker, entity, kind) from filename; None if not a ledger. Applies Dhan fix."""
    b = os.path.basename(path); low = b.lower()
    ent = ("DHR" if b.startswith("DHR") else "HHR" if b.startswith("HHR")
           else "SDR" if b.startswith("SDR") else "Rajani Corp" if b.startswith("Rajani Corp")
           else "HHR" if b.startswith("YourStatement_H53686563") else None)
    if "dhan_ledger" in low:
        broker, kind = "dhan", "ledger"
        if "***REMOVED***" in b or "rajani" in low:   # mislabel: this Dhan account is HHR
            ent = "HHR"
    elif "ledger-report" in low and low.endswith(".xls"):
        broker, kind = "dhan", "ledger"          # Dhan "Ledger Statement" .xls export
    elif "ledger-" in low and (".csv" in low or low.endswith(".xlsx")):
        broker, kind = "zerodha", "ledger"       # Zerodha ledger (CSV or .xlsx)
    elif "yourstatement" in low and low.endswith(".xlsx"):
        broker, kind = "angel_one", "ledger"
    elif "vested_transactions" in low:
        broker, kind = "vested", "income"      # dividends only
    else:
        return None
    return broker, ent, kind


def get_ids(cur, ent_code, broker):
    cur.execute("SELECT id FROM entity WHERE entity_name=%s", (ent_code,))
    e = cur.fetchone()
    if not e:
        return None, None
    cur.execute("""SELECT id FROM account WHERE entity_id=%s AND account_type='BROKER' AND account_name=%s""",
                (e["id"], broker))
    a = cur.fetchone()
    return e["id"], (a["id"] if a else None)


def import_ledger(cur, broker, ent_id, account_id, path, commit, stats):
    closing = {}   # date -> last balance
    cf = 0
    for e in LEDGER_PARSERS[broker](path):
        closing[e["date"]] = e["balance"]
        ftype = classify(e["text"])
        if ftype in ("DEPOSIT", "WITHDRAWAL"):
            amt = -e["credit"] if ftype == "DEPOSIT" else e["debit"]
            if amt == 0:
                continue
            sref = f"{broker}:{ent_id}:{e['date']}:{ftype}:{abs(amt):.0f}:" + \
                   hashlib.md5(e["text"].encode()).hexdigest()[:8]
            if _insert_cf(cur, ent_id, broker, e["date"], ftype, None, amt, "INR", sref, e["text"], commit):
                cf += 1
    # daily closing balances → cash_ledger
    bal = 0
    for d in sorted(closing):
        bal += 1
        if commit:
            cur.execute("""
                INSERT INTO cash_ledger (entity_id, account_id, balance_date, balance, currency, fx_rate, balance_inr, source, created_at)
                VALUES (%s,%s,%s,%s,'INR',1,%s,%s,NOW())
                ON CONFLICT (entity_id, account_id, balance_date)
                DO UPDATE SET balance=EXCLUDED.balance, balance_inr=EXCLUDED.balance_inr, source=EXCLUDED.source
            """, (ent_id, account_id, d, closing[d], closing[d], broker))
    stats["balances"] += bal; stats["cashflows"] += cf
    print(f"  {broker}/{ent_id}: {bal} daily balances, {cf} deposits/withdrawals")


def _insert_cf(cur, ent_id, broker, d, ftype, symbol, amt, ccy, sref, notes, commit):
    if not commit:
        return True
    cur.execute("""INSERT INTO external_cashflow
        (entity_id, broker, flow_date, flow_type, symbol, amount_native, currency, source, source_ref, notes)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (source_ref) DO NOTHING""",
        (ent_id, broker, d, ftype, symbol, round(amt, 2), ccy, broker, sref, (notes or "")[:200]))
    return cur.rowcount > 0


def import_vested_income(cur, ent_id, path, commit, stats):
    n = 0
    for inc in parse_vested_income(path):
        sref = f"vested:{ent_id}:{inc['date']}:{inc['symbol']}:{inc['activity']}:{inc['amount']:.2f}"
        if _insert_cf(cur, ent_id, "vested", inc["date"], inc["flow_type"],
                      inc["symbol"], inc["amount"], "USD", sref, inc["activity"], commit):
            n += 1
    stats["dividends"] += n
    print(f"  vested/{ent_id}: {n} dividend/tax rows (USD)")


def main():
    ap = argparse.ArgumentParser(description="Import broker cash ledgers + dividends.")
    ap.add_argument("--commit", action="store_true")
    ap.add_argument("--dir", default=INCOMING)
    ap.add_argument("--entity", help="force entity for all files (when the filename omits it)")
    args = ap.parse_args()

    conn = psycopg2.connect(host=os.getenv("DB_HOST", "localhost"),
                            database=os.getenv("DB_NAME", "mis_portal"),
                            user=os.getenv("DB_USER", "postgres"),
                            password=os.getenv("DB_PASS") or os.getenv("DB_PASSWORD", ""),
                            cursor_factory=RealDictCursor)
    cur = conn.cursor()
    stats = {"balances": 0, "cashflows": 0, "dividends": 0}
    print(f"{'COMMIT' if args.commit else 'DRY-RUN'}\n")
    for path in sorted(glob.glob(os.path.join(args.dir, "*"))):
        det = detect(path)
        if not det:
            continue
        broker, ent_code, kind = det
        if args.entity:
            ent_code = args.entity          # override: these filenames omit the entity
        ent_id, account_id = get_ids(cur, ent_code, broker)
        if not ent_id:
            print(f"  !! {os.path.basename(path)} — unknown entity {ent_code}"); continue
        print(f"[{broker}/{ent_code}/{kind}] {os.path.basename(path)}")
        try:
            if kind == "ledger":
                import_ledger(cur, broker, ent_id, account_id, path, args.commit, stats)
            elif kind == "income":
                import_vested_income(cur, ent_id, path, args.commit, stats)
        except Exception as ex:
            print(f"  ERROR: {ex}")
    if args.commit:
        conn.commit()
    print(f"\n{stats}")
    print("committed." if args.commit else "dry-run — nothing written.")
    cur.close(); conn.close()


if __name__ == "__main__":
    main()
