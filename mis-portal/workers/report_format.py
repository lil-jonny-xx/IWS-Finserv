#!/usr/bin/env python3
"""
Style vocabulary for the MIS workbook, extracted cell-by-cell from the client's
master format file (MIS-CURRENT-FORMAT.xlsx).

Every fill, font, border and number format below was read off that workbook rather
than invented, so a sheet built through these helpers is visually indistinguishable
from the client's own. Nothing here touches data — see report_generator.py for that.

Palette, by the role it plays in the client's sheets:

  BF9000  dark gold   — weekly-report banners, column headers, grand-total rows
  FFD965  light gold  — weekly-report sub-total rows (e.g. "MF-LIQUID FUND TOTAL")
  F1C232  amber       — Equity Daily Print / All Assets column headers
  FFE599  pale gold   — Equity Daily Print block titles
  FEF2CB  cream       — All Assets sub-total rows
  DEE5EB  pale blue   — Equity Daily Print "Grand Total" rows
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── fills ─────────────────────────────────────────────────────────────────────
GOLD_DARK  = PatternFill("solid", fgColor="FFBF9000")
GOLD_LIGHT = PatternFill("solid", fgColor="FFFFD965")
AMBER      = PatternFill("solid", fgColor="FFF1C232")
GOLD_PALE  = PatternFill("solid", fgColor="FFFFE599")
CREAM      = PatternFill("solid", fgColor="FFFEF2CB")
PALE_BLUE  = PatternFill("solid", fgColor="FFDEE5EB")
NO_FILL    = PatternFill(fill_type=None)

# ── fonts ─────────────────────────────────────────────────────────────────────
F_TITLE     = Font(name="Calibri", size=22, bold=True, color="FF000000")
F_TITLE16   = Font(name="Calibri", size=16, bold=True, color="FF000000")
F_BANNER    = Font(name="Calibri", size=18, bold=True, color="FFFFFFFF")
F_BANNER20  = Font(name="Calibri", size=20, bold=True, color="FFFFFFFF")
F_HDR       = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
F_HDR12     = Font(name="Calibri", size=12, bold=True, color="FFFFFFFF")
F_HDR7      = Font(name="Calibri", size=7,  bold=True, color="FFFFFFFF")
F_HDR_DARK  = Font(name="Calibri", size=11, bold=True, color="FF000000")
F_SUBTOTAL  = Font(name="Calibri", size=12, bold=True, color="FF000000")
F_GRANDTOT  = Font(name="Calibri", size=12, bold=True, color="FFFFFFFF")
F_SECTION   = Font(name="Calibri", size=14, bold=True, color="FF000000")
F_BODY      = Font(name="Calibri", size=11, color="FF000000")
F_BODY12    = Font(name="Calibri", size=12, color="FF000000")
F_BODY_BOLD = Font(name="Calibri", size=12, bold=True, color="FF000000")
F_DATE_SM   = Font(name="Calibri", size=8,  color="FF000000")
F_NOTE      = Font(name="Calibri", size=8,  italic=True, color="FF000000")
F_REMARK    = Font(name="Calibri", size=8,  italic=True, color="FF000000")
F_REMARK_W  = Font(name="Calibri", size=8,  italic=True, color="FFFFFFFF")

# ── borders ───────────────────────────────────────────────────────────────────
_THIN   = Side(style="thin",   color="FF000000")
_MED    = Side(style="medium", color="FF000000")
_DOUBLE = Side(style="double", color="FF000000")

B_THIN      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
B_BOX       = Border(left=_MED,  right=_MED,  top=_MED,  bottom=_MED)
B_BAND      = Border(left=_THIN, right=_THIN, top=_MED,  bottom=_MED)
B_TOTAL     = Border(left=_THIN, right=_THIN, top=_DOUBLE, bottom=_MED)
B_HDR       = Border(left=_THIN, right=_THIN, top=_MED,  bottom=_MED)
B_NONE      = Border()

# ── number formats ────────────────────────────────────────────────────────────
MONEY     = r'_ * #,##0_ ;_ * \-#,##0_ ;_ * "-"??_ ;_ @_ '
MONEY_TOT = r'_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)'
PCT2      = '0.00%'
PCT0      = '0%'
PCT1      = '0.0%'
QTY       = '#,##0'
PRICE     = '#,##0.00'
DATE_LONG = r'd\ mmm\ yyyy'
DATE_SHORT= 'D/M/YYYY'
DATE_DMY  = 'd-mmm-yy'

# Alignments reused everywhere.
AL_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_CENTER      = Alignment(horizontal="center", vertical="center")
AL_LEFT        = Alignment(horizontal="left",   vertical="center")
AL_LEFT_WRAP   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
AL_RIGHT       = Alignment(horizontal="right",  vertical="center")


# ── primitives ────────────────────────────────────────────────────────────────

def style_cell(ws, row, col, value=None, *, font=None, fill=None, border=None,
               align=None, fmt=None):
    """Write one cell and stamp its full style. Returns the cell."""
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    if font   is not None: c.font = font
    if fill   is not None: c.fill = fill
    if border is not None: c.border = border
    if align  is not None: c.alignment = align
    if fmt    is not None: c.number_format = fmt
    return c


def paint_row(ws, row, first_col, last_col, *, font=None, fill=None, border=None,
              align=None):
    """Stamp a style across a whole row span without touching the values.

    Used to carry a banner's fill to the end of its block so a merged title reads as
    one continuous bar rather than a coloured cell followed by white space — the way
    the client's own sheets look.
    """
    for c in range(first_col, last_col + 1):
        style_cell(ws, row, c, font=font, fill=fill, border=border, align=align)


def banner(ws, row, first_col, last_col, text, *, font=None, fill=None,
           border=None, align=None, height=None):
    """Merged, filled title bar spanning first_col..last_col."""
    paint_row(ws, row, first_col, last_col,
              font=font or F_BANNER, fill=fill or GOLD_DARK,
              border=border, align=align or AL_CENTER)
    ws.merge_cells(start_row=row, start_column=first_col,
                   end_row=row, end_column=last_col)
    ws.cell(row=row, column=first_col).value = text
    if height:
        ws.row_dimensions[row].height = height
    return row + 1


def set_widths(ws, widths: dict):
    """widths: {'A': 39.14, ...} — the client's own column widths."""
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


def note(ws, row, col, text, span=None):
    """Small italic footnote, exactly the 8pt italic the client uses."""
    c = style_cell(ws, row, col, text, font=F_NOTE, align=AL_LEFT_WRAP)
    if span:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=span)
    return row + 1
