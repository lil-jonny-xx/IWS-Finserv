#!/usr/bin/env python3
"""
Report generator — produces per-entity and combined portfolio Excel reports.
Call generate_reports(conn, generated_by_user_id) to create reports.
"""
import os
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional
from collections import defaultdict, deque

# Project root on sys.path so `workers.` / `equity.` resolve when this file is run
# DIRECTLY (cron_wrapper spawns `python workers/report_generator.py` as a fresh
# subprocess, which has neither the wrapper's sys.path nor a package context).
# Without this the scheduled run dies on ModuleNotFoundError before it starts.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Split/bonus lot adjustment. Imported lazily-safe: this module is also used in
# contexts where the corporate_action table may not exist yet, and load_actions()
# returns {} in that case rather than raising.
from workers.corporate_actions import (  # noqa: E402
    load_actions as _ca_load, apply_actions as _ca_apply,
)
# Shared gold/silver/commodity classifier — the same one that tags equity_holding
# and drives the Gold/Silver page, so realised gains bucket the way holdings do.
from equity.asset_class import (  # noqa: E402
    classify_asset_class as _ac_classify, load_overrides as _ac_overrides,
    GOLD_SILVER_COMMODITY as _AC_COMMODITY,
)

# Sheet layout. Every block, header and fill in these builders was taken off the
# client's master format workbook, so the generated file matches it cell for cell.
from workers import report_format as _rf  # noqa: E402
from workers.report_sheets import (  # noqa: E402
    classify_bundle, build_weekly_report, build_realised_pnl,
    build_equity_daily_print as _render_equity_daily_print,
    build_all_assets_daily, build_all_entities_weekly,
)

REPORTS_DIR = "/var/www/mis-portal/reports"

# Sheet layout — including the colour palette and every column position — now lives
# in workers/report_format.py and workers/report_sheets.py, taken cell for cell off
# the client's master format workbook. This module owns the DATA only.

# ── DB helpers ────────────────────────────────────────────────────────────────

def _fetch_entities(conn):
    cur = conn.cursor()
    cur.execute("SELECT id, entity_name FROM entity ORDER BY id")
    rows = cur.fetchall()
    cur.close()
    return rows


def _fetch_equity_daily_data(conn, entity_ids: list):
    """Fetch equity holdings with quantity, avg_cost, current_price for Equity Daily Print."""
    cur = conn.cursor()
    placeholders = ",".join(["%s"] * len(entity_ids))
    cur.execute(f"""
        SELECT
            eh.entity_id, e.entity_name,
            COALESCE(eh.symbol_override, eh.symbol) AS symbol,
            sm.security_name,
            eh.isin, eh.broker,
            eh.quantity, eh.avg_cost,
            eh.cost, eh.current_price, eh.current_market_value,
            eh.pnl_inception, eh.returns_inception_pct,
            eh.first_invested_date
        FROM equity_holding eh
        JOIN entity e ON e.id = eh.entity_id
        LEFT JOIN security_master sm ON sm.isin = eh.isin
        WHERE eh.entity_id IN ({placeholders})
          AND COALESCE(eh.asset_class, 'equity') NOT IN ('gold', 'silver', 'commodity')
        ORDER BY eh.entity_id, eh.current_market_value DESC NULLS LAST
    """, entity_ids)
    rows = cur.fetchall()
    cur.close()
    return rows


def _merge_edp_rows(rows: list, cross_entity: bool = False) -> list:
    """
    Merge same-ISIN holdings. If cross_entity=False, merges within each entity.
    If cross_entity=True, merges across all entities (for grand total).
    Returns list sorted by current_market_value desc.
    """
    groups: dict = defaultdict(list)
    for r in rows:
        isin = (r.get("isin") or "").strip()
        sym  = r.get("symbol") or ""
        if cross_entity:
            key = isin if isin else sym
        else:
            key = (r["entity_id"], isin if isin else sym)
        groups[key].append(r)

    merged = []
    for key, rrows in groups.items():
        if len(rrows) == 1:
            merged.append(dict(rrows[0]))
            continue

        def _fsum(k):
            return sum(float(r[k]) for r in rrows if r.get(k) is not None)

        qty     = _fsum("quantity")
        cost    = _fsum("cost")
        cmv     = _fsum("current_market_value")
        pnl     = _fsum("pnl_inception")
        avg_c   = cost / qty if qty else None
        ret_inc = pnl / cost * 100 if cost else None
        cur_p   = next((float(r["current_price"]) for r in rrows if r.get("current_price")), None)

        m = dict(rrows[0])
        m.update({
            "quantity":             qty,
            "avg_cost":             avg_c,
            "cost":                 cost,
            "current_price":        cur_p,
            "current_market_value": cmv,
            "pnl_inception":        pnl,
            "returns_inception_pct": ret_inc,
        })
        merged.append(m)

    return sorted(merged, key=lambda r: -(float(r.get("current_market_value") or 0)))


# ---------------------------------------------------------------------------
# The "Mkt Value 31-Mar" column
#
# `holding.market_value_as_on` / `equity_holding.market_value_as_on` is NOT a 31-March
# anchor despite the column header: every writer sets it to the CURRENT market value
# (equity/models.py — "same as current_market_value; date stored separately"), so it
# was identical to the Current Mkt Value column on 100% of rows in both tables. Two
# columns, one number, one of them lying about which date it belongs to.
#
# So the reports derive the real thing here instead, from the same 31-March anchors the
# FY-growth columns use (security_price_history / nav_history, backfilled by
# fy_price_backfill). Where no anchor exists — an SME board Yahoo doesn't carry, a
# holding bought after 31-March, a manual entry with no price series — the cell is left
# BLANK rather than back-filled with today's value, which is the whole point.
# ---------------------------------------------------------------------------

def _feed_staleness(conn, as_of: date, entity_id: Optional[int] = None) -> list[str]:
    """Per-feed 'as of' where it is BEHIND the report date, as ['IBKR 05-Aug (4d)', …].

    Every sheet is stamped "As on <today>", but the feeds behind it do not all reach
    today: a throttled IBKR Flex pull serves a days-old statement, DBS is a weekly CSV
    someone has to upload, and a dead token freezes an entity indefinitely. Stamping the
    report date over all of that is how a 26-day-old position came to read as current.
    So the sheets now carry the real vintage of anything lagging, and say nothing when
    every feed is current.

    Reads as_of_date, which is the one column that means "the data is from this date"
    (updated_at only means "a worker touched the row").
    """
    # Measured against the last TRADING day, not the calendar date: a report run on a
    # Sunday finds every market feed showing Friday, which is current, not stale. Only
    # a feed that missed the last session is worth a warning.
    last_session = as_of
    while last_session.weekday() >= 5:            # 5=Sat, 6=Sun
        last_session -= timedelta(days=1)
    scope = " AND entity_id = %(eid)s" if entity_id else ""
    params = {"as_of": last_session, "eid": entity_id}
    cur = conn.cursor()
    out: list[str] = []
    try:
        # MIN, not MAX, at every level: this is a warning, so a feed is only as fresh as
        # its stalest entity. Taking the max let DHR's current IBKR pull hide SDR's,
        # which a dead Flex token had frozen 26 days earlier — exactly the kind of
        # masking this note exists to prevent.
        cur.execute(f"""
            SELECT label, MIN(d) AS d FROM (
                SELECT UPPER(broker) AS label, MIN(as_of_date) AS d
                  FROM foreign_equity_holding WHERE as_of_date IS NOT NULL {scope}
                 GROUP BY broker
                UNION ALL
                SELECT UPPER(broker), MIN(as_of_date)
                  FROM equity_holding WHERE as_of_date IS NOT NULL {scope}
                 GROUP BY broker
                UNION ALL
                SELECT 'MUTUAL FUNDS', MIN(as_of_date)
                  FROM holding WHERE as_of_date IS NOT NULL {scope}
            ) f
            WHERE d < %(as_of)s
            GROUP BY label ORDER BY MIN(d), label
        """, params)
        for r in cur.fetchall():
            lag = (as_of - r["d"]).days       # reported against the real report date
            out.append(f"{r['label']} {r['d'].strftime('%d-%b')} ({lag}d)")
    except Exception:
        conn.rollback()
    finally:
        cur.close()
    return out


def _staleness_note(conn, as_of: date, entity_id: Optional[int] = None) -> str:
    """One-line warning naming any feed that has not refreshed to the report date;
    empty string when every feed is current. Rendered as a footnote at the foot of
    the weekly page (the client's format has no dedicated cell for it)."""
    stale = _feed_staleness(conn, as_of, entity_id)
    return ("*DATA BEHIND THIS DATE: " + ", ".join(stale)) if stale else ""


def _fy_mar31_for(as_of: Optional[date] = None) -> date:
    """31-Mar opening the current FY. Defined here (not via _fy_mar31, declared later)
    so the fetch helpers can default it without a forward reference."""
    d = as_of or date.today()
    return date(d.year if d.month >= 4 else d.year - 1, 3, 31)


def _fetch_mf_holdings(conn, entity_id: Optional[int] = None):
    """Fetch MF holdings. DB security_type values: MF_DEBT, MF_EQUITY, MF_HYBRID."""
    cur = conn.cursor()
    mar31 = _fy_mar31_for()
    q = """
        SELECT
            h.entity_id, e.entity_name,
            sm.security_name, sm.asset_class, sm.security_type,
            h.invested_amount   AS cost,
            h.current_value,
            h.prev_week_value,
            -- Units held now, valued at the FY-opening NAV. 31-Mar is routinely a
            -- holiday, so take the last NAV in a short window before it.
            (SELECT nh.nav * h.quantity FROM nav_history nh
              WHERE nh.security_id = h.security_id
                AND nh.nav_date <= %(mar31)s AND nh.nav_date >= %(mar31)s - 10
              ORDER BY nh.nav_date DESC LIMIT 1)   AS market_value_as_on,
            h.pnl_ytd, h.pnl_inception,
            h.returns_ytd_pct, h.returns_inception_pct, h.cagr_inception_pct,
            h.xirr_inception_pct,
            h.first_invested_date,
            h.weekly_change, h.exposure_pct, h.remarks
        FROM holding h
        JOIN entity e  ON e.id  = h.entity_id
        JOIN security_master sm ON sm.id = h.security_id
        {where}
        ORDER BY sm.asset_class, sm.security_name
    """
    # quantity > 0 — a with-zero-balance CAS carries every fully-exited scheme as a
    # zero-quantity holding (332 of them across the book) purely so the closed-folio
    # history reaches mf_transaction for realised gains. They are not positions and
    # must not be listed as holdings in the report.
    if entity_id:
        cur.execute(q.format(where="WHERE h.quantity > 0 AND h.entity_id = %(eid)s"),
                    {"mar31": mar31, "eid": entity_id})
    else:
        cur.execute(q.format(where="WHERE h.quantity > 0"), {"mar31": mar31})
    rows = cur.fetchall()
    cur.close()
    return rows


def _fetch_equity_holdings(conn, entity_id: Optional[int] = None):
    """Domestic direct-equity holdings (equity_holding). Foreign brokers live in
    foreign_equity_holding — see _fetch_foreign_equity_holdings — and are reported
    in their own section so totals stay whole without mixing the two."""
    cur = conn.cursor()
    mar31 = _fy_mar31_for()
    q = """
        SELECT
            eh.entity_id, e.entity_name,
            eh.broker, COALESCE(eh.symbol_override, eh.symbol) AS symbol, eh.isin,
            eh.cost, eh.current_market_value AS current_value,
            eh.prev_week_value,
            -- Shares held now, valued at the FY-opening close (see the note above
            -- _fetch_mf_holdings). security_symbol_map resolves the broker ticker to
            -- the Yahoo one the anchors are stored under.
            (SELECT sph.close * eh.quantity
               FROM security_symbol_map m
               JOIN security_price_history sph ON sph.yahoo_symbol = m.resolved_symbol
              WHERE m.symbol = eh.symbol AND sph.price_date = %(mar31)s) AS market_value_as_on,
            eh.pnl_ytd, eh.pnl_inception,
            eh.returns_ytd_pct, eh.returns_inception_pct, eh.cagr_inception_pct,
            eh.first_invested_date,
            eh.weekly_change, eh.exposure_pct, eh.remarks
        FROM equity_holding eh
        JOIN entity e ON e.id = eh.entity_id
        WHERE COALESCE(eh.asset_class, 'equity') NOT IN ('gold', 'silver', 'commodity')
        {ent}
        ORDER BY eh.symbol
    """
    if entity_id:
        cur.execute(q.format(ent="AND eh.entity_id = %(eid)s"),
                    {"mar31": mar31, "eid": entity_id})
    else:
        cur.execute(q.format(ent=""), {"mar31": mar31})
    rows = cur.fetchall()
    cur.close()
    return rows


def _fetch_foreign_equity_holdings(conn, entity_id: Optional[int] = None):
    """Foreign (multi-currency) direct-equity holdings (foreign_equity_holding).
    Returns the INR-converted columns used by the by-broker rollup AND the native
    currency columns used by the Foreign Equity Print detail sheet."""
    cur = conn.cursor()
    mar31 = _fy_mar31_for()
    q = """
        SELECT
            eh.entity_id, e.entity_name,
            eh.broker, COALESCE(eh.symbol_override, eh.symbol) AS symbol, eh.isin, eh.exchange,
            eh.cost, eh.current_market_value AS current_value,
            eh.prev_week_value,
            -- Shares held now at the FY-opening close, converted at the FY-opening FX
            -- rate — the column is in ₹, so a year's currency move belongs in it.
            -- symbol_override IS the Yahoo ticker for foreign rows (a bare LSE/SIX
            -- ticker collides with a US fund), so no symbol map is involved.
            (SELECT sph.close * eh.quantity * COALESCE(fx.rate, 1)
               FROM security_price_history sph
               LEFT JOIN LATERAL (
                    SELECT r.rate FROM fx_rate r
                     WHERE r.from_currency = COALESCE(eh.currency, 'USD')
                       AND r.to_currency = 'INR'
                       AND r.rate_date <= %(mar31)s AND r.rate_date >= %(mar31)s - 10
                     ORDER BY r.rate_date DESC LIMIT 1) fx ON TRUE
              WHERE sph.yahoo_symbol = COALESCE(NULLIF(eh.symbol_override,''), eh.symbol)
                AND sph.price_date = %(mar31)s)      AS market_value_as_on,
            eh.pnl_ytd, eh.pnl_inception,
            eh.returns_ytd_pct, eh.returns_inception_pct, eh.cagr_inception_pct,
            eh.first_invested_date,
            eh.weekly_change, eh.exposure_pct, eh.remarks,
            eh.currency, eh.fx_rate, eh.quantity,
            eh.avg_cost_native, eh.cost_native,
            eh.current_price_native, eh.current_market_value_native,
            eh.xirr_inception_pct
        FROM foreign_equity_holding eh
        JOIN entity e ON e.id = eh.entity_id
        WHERE COALESCE(eh.asset_class, 'equity') NOT IN ('gold', 'silver', 'commodity')
        {ent}
        ORDER BY eh.broker, eh.symbol
    """
    if entity_id:
        cur.execute(q.format(ent="AND eh.entity_id = %(eid)s"),
                    {"mar31": mar31, "eid": entity_id})
    else:
        cur.execute(q.format(ent=""), {"mar31": mar31})
    rows = cur.fetchall()
    cur.close()
    return rows


def _fetch_commodity_holdings(conn, entity_id: Optional[int] = None):
    """Gold / silver / commodity holdings (asset_class in gold/silver/commodity),
    unioned across domestic (equity_holding) and foreign (foreign_equity_holding).
    These are split out of the Equity / Foreign Equity sections — mirroring the
    dedicated Commodities page — and reported on their own. Foreign rows use their
    INR-converted value columns so the grand totals stay whole."""
    cur = conn.cursor()
    mar31 = _fy_mar31_for()
    # FY-opening value (see the note above _fetch_mf_holdings). Domestic rows resolve
    # their Yahoo ticker through security_symbol_map; foreign rows carry it directly in
    # symbol_override — hence the two spellings of the same subquery.
    mv_dom = """
            (SELECT sph.close * eh.quantity
               FROM security_symbol_map m
               JOIN security_price_history sph ON sph.yahoo_symbol = m.resolved_symbol
              WHERE m.symbol = eh.symbol AND sph.price_date = %(mar31)s)
    """
    mv_fgn = """
            (SELECT sph.close * eh.quantity * COALESCE(fx.rate, 1)
               FROM security_price_history sph
               LEFT JOIN LATERAL (
                    SELECT r.rate FROM fx_rate r
                     WHERE r.from_currency = COALESCE(eh.currency, 'USD')
                       AND r.to_currency = 'INR'
                       AND r.rate_date <= %(mar31)s AND r.rate_date >= %(mar31)s - 10
                     ORDER BY r.rate_date DESC LIMIT 1) fx ON TRUE
              WHERE sph.yahoo_symbol = COALESCE(NULLIF(eh.symbol_override,''), eh.symbol)
                AND sph.price_date = %(mar31)s)
    """

    def cols(mv):
        return f"""
            eh.entity_id, e.entity_name, eh.broker,
            COALESCE(eh.symbol_override, eh.symbol) AS symbol, eh.isin,
            COALESCE(eh.asset_class, 'commodity') AS asset_class,
            eh.cost, eh.current_market_value AS current_value,
            eh.prev_week_value, {mv} AS market_value_as_on,
            eh.pnl_ytd, eh.pnl_inception,
            eh.returns_ytd_pct, eh.returns_inception_pct, eh.cagr_inception_pct,
            eh.first_invested_date, eh.weekly_change, eh.exposure_pct, eh.remarks
        """

    ent = "AND eh.entity_id = %(eid)s" if entity_id else ""
    q = f"""
        SELECT {cols(mv_dom)} FROM equity_holding eh JOIN entity e ON e.id = eh.entity_id
        WHERE COALESCE(eh.asset_class, 'equity') IN ('gold', 'silver', 'commodity') {ent}
        UNION ALL
        SELECT {cols(mv_fgn)} FROM foreign_equity_holding eh JOIN entity e ON e.id = eh.entity_id
        WHERE COALESCE(eh.asset_class, 'equity') IN ('gold', 'silver', 'commodity') {ent}
        ORDER BY broker, symbol
    """
    params = {"mar31": mar31}
    if entity_id:
        params["eid"] = entity_id
    cur.execute(q, params)
    rows = cur.fetchall()
    cur.close()
    return rows


def _fetch_manual_inputs(conn, entity_id: Optional[int] = None):
    """Latest manual input per (entity_id, category, label)."""
    cur = conn.cursor()
    q = """
        SELECT DISTINCT ON (entity_id, category, label)
            entity_id, category, label,
            cost, current_value, prev_week_value,
            currency, raw_amount, fx_rate,
            inception_date, notes, updated_at
        FROM manual_input
        {where}
        ORDER BY entity_id, category, label, updated_at DESC
    """
    if entity_id:
        cur.execute(q.format(where="WHERE entity_id = %s"), (entity_id,))
    else:
        cur.execute(q.format(where=""))
    rows = cur.fetchall()
    cur.close()
    return rows


# ── equity merge (mirrors frontend Combined view) ─────────────────────────────

def _merge_equity_by_symbol(eq_rows: list) -> list:
    """
    Collapse same symbol + same entity rows across brokers into one row.
    Qty/cost/CMV/P&L summed; returns%, CAGR recalculated; earliest inception date kept.
    broker field becomes a comma-joined list of all brokers involved.
    """
    groups: dict = defaultdict(list)
    for h in eq_rows:
        isin = (h.get("isin") or "").strip()
        key  = (h["entity_id"], isin if isin else h["symbol"])
        groups[key].append(h)

    merged = []
    for (entity_id, symbol), rows in groups.items():
        if len(rows) == 1:
            merged.append(dict(rows[0]))
            continue

        def _sum(key):
            vals = [float(r[key]) for r in rows if r.get(key) is not None]
            return sum(vals) if vals else None

        cost    = _sum("cost")
        cur_val = _sum("current_value")
        pnl_inc = _sum("pnl_inception")
        pnl_ytd = _sum("pnl_ytd")
        prev_wk = _sum("prev_week_value")
        wkly    = _sum("weekly_change")
        mkt_mar = _sum("market_value_as_on")

        returns_inc = (pnl_inc / cost * 100) if cost and pnl_inc is not None else None
        returns_ytd = (pnl_ytd / cost * 100) if cost and pnl_ytd is not None else None

        dates = [r["first_invested_date"] for r in rows if r.get("first_invested_date")]
        first_date = min(dates) if dates else None
        cagr = None
        if first_date and cost and cur_val and cost > 0 and cur_val > 0:
            years = (date.today() - first_date).days / 365.25
            if years >= 1.0:
                try:
                    cagr = ((cur_val / cost) ** (1 / years) - 1) * 100
                except Exception:
                    pass

        _BROKER_LABELS = {"zerodha": "Zerodha", "angel_one": "Angel One", "dhan": "Dhan"}
        brokers_str  = ", ".join(sorted({_BROKER_LABELS.get(r["broker"], r["broker"].title()) for r in rows}))
        display_sym  = next((r["symbol"] for r in rows if r.get("symbol")), rows[0].get("isin", ""))
        merged_row = dict(rows[0])
        merged_row.update({
            "symbol":               display_sym,
            "cost":                 cost,
            "current_value":        cur_val,
            "prev_week_value":      prev_wk,
            "weekly_change":        wkly,
            "pnl_inception":        pnl_inc,
            "pnl_ytd":              pnl_ytd,
            "returns_inception_pct": returns_inc,
            "returns_ytd_pct":       returns_ytd,
            "cagr_inception_pct":   cagr,
            "first_invested_date":  first_date,
            "market_value_as_on":   mkt_mar,
            "broker":               brokers_str,
        })
        merged.append(merged_row)

    return sorted(merged, key=lambda h: h["symbol"])


_BROKER_DISPLAY = {"zerodha": "Zerodha", "angel_one": "Angel One", "dhan": "Dhan",
                   "ibkr": "Interactive Brokers", "vested": "Vested", "dbs": "DBS Wealth"}
_PMS_SOURCE_DISPLAY = {"nuvama_pms": "Nuvama", "zerodha_pms": "Zerodha PMS"}
def _fetch_pms_holdings(conn, entity_id: Optional[int] = None):
    """PMS positions (pms_holding) — Nuvama + Zerodha-PMS, equity and cash."""
    cur = conn.cursor()
    q = """
        SELECT entity_id, source, holding_type,
               COALESCE(cost, 0) AS cost, COALESCE(market_value, 0) AS market_value
        FROM pms_holding
        {where}
    """
    if entity_id:
        cur.execute(q.format(where="WHERE entity_id = %s"), (entity_id,))
    else:
        cur.execute(q.format(where=""))
    rows = cur.fetchall()
    cur.close()
    return rows


def _pms_report_rows(pms_rows: list) -> list:
    """
    Shape pms_holding into report rows: per source, one summed Equity row
    (blended return) and one Cash row. Not per-stock — matching the summary
    style of the All Assets equity rollup.
    """
    agg: dict = defaultdict(lambda: {"eq_cost": 0.0, "eq_mv": 0.0, "cash": 0.0})
    for r in pms_rows:
        a = agg[r["source"]]
        if r["holding_type"] == "cash":
            a["cash"] += float(r["market_value"])
        else:
            a["eq_cost"] += float(r["cost"]); a["eq_mv"] += float(r["market_value"])

    out = []
    for source in sorted(agg):
        disp, a = _PMS_SOURCE_DISPLAY.get(source, source), agg[source]
        if a["eq_mv"] or a["eq_cost"]:
            ret = ((a["eq_mv"] - a["eq_cost"]) / a["eq_cost"] * 100) if a["eq_cost"] else None
            out.append({
                "label": f"PMS — {disp} (Equity)",
                "cost": a["eq_cost"], "current_value": a["eq_mv"],
                "returns_inception_pct": ret,
            })
        if a["cash"]:
            out.append({
                "label": f"PMS — {disp} (Cash)",
                "cost": a["cash"], "current_value": a["cash"],
            })
    return out


def _fetch_broker_cash(conn, entity_id: Optional[int] = None):
    """Available cash per (entity, broker) from broker_cash."""
    cur = conn.cursor()
    q = """
        SELECT entity_id, broker, COALESCE(balance, 0) AS balance
        FROM broker_cash
        {where}
    """
    if entity_id:
        cur.execute(q.format(where="WHERE entity_id = %s"), (entity_id,))
    else:
        cur.execute(q.format(where=""))
    rows = cur.fetchall()
    cur.close()
    return rows


def _broker_cash_report_rows(cash_rows: list) -> list:
    """One report row per broker cash balance (cost == value, no P&L)."""
    out = []
    for r in sorted(cash_rows, key=lambda x: x["broker"]):
        bal = float(r["balance"])
        out.append({
            "label": f"{_BROKER_DISPLAY.get(r['broker'], r['broker'].title())} (Cash)",
            "cost": bal, "current_value": bal,
        })
    return out


def _fetch_bank_accounts(conn, entity_id: Optional[int] = None):
    """Bank-account cash per (entity, bank), in native currency, from bank_account."""
    cur = conn.cursor()
    q = """
        SELECT entity_id, bank_name, account_type, currency, COALESCE(balance, 0) AS balance
        FROM bank_account
        {where}
    """
    if entity_id:
        cur.execute(q.format(where="WHERE entity_id = %s"), (entity_id,))
    else:
        cur.execute(q.format(where=""))
    rows = cur.fetchall()
    cur.close()
    return rows


def _bank_account_report_rows(conn, bank_rows: list, as_of: date) -> list:
    """One report row per bank account, native balance converted to INR at the
    as-of fx rate (cost == value, no P&L). A balance whose currency has no rate
    is skipped — it can't be valued in INR (matches the /overview behaviour)."""
    out = []
    for r in sorted(bank_rows, key=lambda x: (x["bank_name"], x["account_type"])):
        fx = _fx_rate_on(conn, r["currency"], as_of)
        if fx is None:
            continue
        bal_inr = float(r["balance"]) * fx
        ccy = "" if r["currency"] == "INR" else f" ({r['currency']})"
        out.append({
            "label": f"{r['bank_name']}{ccy} (Bank)",
            "cost": bal_inr, "current_value": bal_inr,
        })
    return out



# Friendly group label keyed by the group's lead (first) entity code.
# Group display names the client uses on the reports. SDR's PAN group is a
# single-member one, but the client's sheets still title it "Stuti GROUP".
GROUP_DISPLAY = {"DHR": "Dhruv", "HHR": "Harsh", "SDR": "Stuti"}
def _all_pan_groups(conn) -> list[dict]:
    """
    EVERY PAN group → [{label, entity_ids}] in group order.

    Unlike _pan_groups (which skips single-entity groups because they are already
    covered by their own entity sheet), this includes them: the client's All Entities
    monitor gives each PAN group its own column block, and leaving the singletons out
    would make the ALL RAJANI GROUP total disagree with the per-entity pages.
    """
    cur = conn.cursor()
    cur.execute("""
        SELECT pg.id, e.id AS entity_id, e.entity_name
        FROM pan_group pg JOIN entity e ON e.pan_group_id = pg.id
        ORDER BY pg.id, e.id
    """)
    rows = cur.fetchall(); cur.close()
    by_group: dict = defaultdict(list)
    for r in rows:
        by_group[r["id"]].append(r)

    groups = []
    for members in by_group.values():
        lead = members[0]["entity_name"]
        display = GROUP_DISPLAY.get(lead)
        # "<Name> GROUP" where the client has a name for it, else the entity's own
        # code — the client writes "Rajani Corp", not "Rajani Corp GROUP".
        label = f"{display} GROUP" if display else lead
        groups.append({"label": label,
                       "entity_ids": [m["entity_id"] for m in members]})
    return groups


def _fetch_dividends(conn, entity_ids: list, as_of: date) -> list:
    """Dividends credited in the current FY, for the realised-gains page.

    Indian dividends never reach us through the broker feed — they are derived from
    yfinance rates replayed over the ledger (see workers/dividend_worker.py) — so this
    reads the `dividend` table rather than any trade stream.
    """
    if not entity_ids:
        return []
    cur = conn.cursor()
    ph = ",".join(["%s"] * len(entity_ids))
    try:
        cur.execute(f"""
            SELECT d.entity_id, sm.security_name, d.pay_date, d.ex_date,
                   d.amount, d.currency, d.source
            FROM dividend d
            LEFT JOIN security_master sm ON sm.id = d.security_id
            WHERE d.entity_id IN ({ph})
              AND COALESCE(d.pay_date, d.ex_date) >= %s
              AND COALESCE(d.pay_date, d.ex_date) <= %s
            ORDER BY COALESCE(d.pay_date, d.ex_date)
        """, entity_ids + [_fy_start(as_of), as_of])
        rows = cur.fetchall()
    except Exception:
        # The table postdates some deployments; a missing one must not fail the run.
        conn.rollback(); rows = []
    cur.close()
    out = []
    for r in rows:
        out.append({"security_name": r["security_name"] or "",
                    "pay_date": r["pay_date"] or r["ex_date"],
                    "amount": float(r["amount"]) if r["amount"] is not None else None,
                    "broker": (r["source"] or "").upper() or "OTHER"})
    return out


def _edp_holdings(conn, entity_ids_by_code: dict) -> tuple:
    """
    Per-entity domestic and foreign direct-equity holdings for the Equity Daily Print
    tab, merged by ISIN within each entity and sorted highest market value first
    (which is what that tab's own column header promises).

    Returns ({code: [holding, ...]}, {code: [holding, ...]}) — domestic, foreign.
    """
    domestic, foreign = {}, {}
    for code, eid in entity_ids_by_code.items():
        merged = _merge_edp_rows(_fetch_equity_daily_data(conn, [eid]),
                                 cross_entity=True)
        if merged:
            domestic[code] = merged

    for code, eid in entity_ids_by_code.items():
        pr = []
        for r in _fetch_foreign_equity_holdings(conn, eid):
            # The client's foreign blocks are headed "($)", so they show the NATIVE
            # price, cost and market value — not the INR conversion the weekly page
            # uses. P&L has no native column, so it is derived here from the same
            # native pair rather than mixing a ₹ figure into a $ block.
            cost_n = r.get("cost_native")
            cmv_n  = r.get("current_market_value_native")
            pnl_n  = ((float(cmv_n) - float(cost_n))
                      if (cmv_n is not None and cost_n is not None) else None)
            pr.append({**r,
                       "security_name":        r.get("symbol"),
                       "avg_cost":             r.get("avg_cost_native"),
                       "cost":                 cost_n,
                       "current_price":        r.get("current_price_native"),
                       "current_market_value": cmv_n,
                       "pnl_inception":        pnl_n,
                       "returns_inception_pct": ((pnl_n / float(cost_n) * 100)
                                                 if (pnl_n is not None and cost_n)
                                                 else None)})
        pr = _merge_edp_rows(pr, cross_entity=True)
        if pr:
            foreign[code] = pr
    return domestic, foreign


def _bundle_for(conn, entity_ids: list, as_of: date) -> dict:
    """
    Gather + merge holdings for one entity or a set of entities (a group).
    Direct equity is merged by ISIN ACROSS the whole bundle so the same share held
    via several brokers (or several entities in a group) collapses to one line, with
    every broker named.  Returns {mf, eq, fe, comm, pms, cash, bank, manual_by_cat}.
    """
    mf_rows, eq_raw, fe_raw, comm_raw, man_rows = [], [], [], [], []
    pms_raw, cash_raw, bank_raw = [], [], []
    for eid in entity_ids:
        mf_rows.extend(_fetch_mf_holdings(conn, eid))
        eq_raw.extend(_fetch_equity_holdings(conn, eid))
        fe_raw.extend(_fetch_foreign_equity_holdings(conn, eid))
        comm_raw.extend(_fetch_commodity_holdings(conn, eid))
        pms_raw.extend(_fetch_pms_holdings(conn, eid))
        cash_raw.extend(_fetch_broker_cash(conn, eid))
        bank_raw.extend(_fetch_bank_accounts(conn, eid))
        man_rows.extend(_fetch_manual_inputs(conn, eid))

    # Collapse entity_id so _merge_equity_by_symbol merges across the whole bundle.
    for r in eq_raw + fe_raw + comm_raw:
        r["entity_id"] = 0
    eq   = _merge_equity_by_symbol(eq_raw)
    fe   = _merge_equity_by_symbol(fe_raw)    # foreign, merged by symbol across the bundle
    comm = _merge_equity_by_symbol(comm_raw)  # gold/silver/commodity, merged by symbol

    # Normalise broker labels (single-broker rows arrive raw e.g. "zerodha").
    for r in eq + fe + comm:
        b = r.get("broker")
        if b:
            r["broker"] = ", ".join(
                _BROKER_DISPLAY.get(p.strip(), p.strip().title()) for p in str(b).split(","))

    man_by_cat: dict = defaultdict(list)
    for m in man_rows:
        man_by_cat[m["category"]].append(m)
    return {"mf": mf_rows, "eq": eq, "fe": fe, "comm": comm,
            "pms":  _pms_report_rows(pms_raw),
            "cash": _broker_cash_report_rows(cash_raw),
            "bank": _bank_account_report_rows(conn, bank_raw, as_of),
            "manual_by_cat": man_by_cat}
_BENCHMARK_ORDER = ["SENSEX", "NIFTY", "GS2032_YTM", "GS2032_PRICE", "GS2030_YTM", "GS2030_PRICE"]


def _fetch_benchmarks(conn, as_of: date) -> list[dict]:
    """
    Current / previous-week / 31-Mar values per benchmark (derived from the
    market_benchmark history) + week% and YTD% change.  Returns [] if the table
    is absent (migration not yet run) so callers still work.

    NOT used by the workbook any more — the Market Statistics block was dropped from
    the reports. This is now the data source for GET /api/v1/benchmarks (main.py),
    which feeds the dashboard market rail and the top ticker. Deleting it along with
    the report block took the overview page's market data down with it (2026-08-10),
    so it lives on here as an API helper. Move it to a shared module if this file's
    report-only role ever needs to be strict.
    """
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT code, label, as_of_date, value, unit, prev_close
            FROM market_benchmark
            WHERE value IS NOT NULL AND as_of_date <= %s
            ORDER BY code, as_of_date
        """, (as_of,))
        rows = cur.fetchall()
    except Exception:
        conn.rollback(); cur.close()
        return []
    cur.close()

    series: dict = defaultdict(list)
    label_of, unit_of, prev_close_of = {}, {}, {}
    for r in rows:
        series[r["code"]].append((r["as_of_date"], float(r["value"])))
        label_of[r["code"]] = r["label"]
        unit_of[r["code"]] = r["unit"]
        # Rows are ordered by date, so the last one wins — the prior close belonging
        # to the most recent reading, which is the one day% is measured against.
        prev_close_of[r["code"]] = float(r["prev_close"]) if r["prev_close"] is not None else None

    prev_cut = as_of.fromordinal(as_of.toordinal() - 7)
    mar31    = _fy_mar31_for(as_of)

    def _at_or_before(pairs, cutoff):
        val = None
        for d, v in pairs:                       # pairs sorted ascending
            if d <= cutoff:
                val = v
        return val

    out = []
    codes = [c for c in _BENCHMARK_ORDER if c in series] + \
            [c for c in series if c not in _BENCHMARK_ORDER]
    for code in codes:
        pairs = series[code]
        current   = pairs[-1][1]
        prev_week = _at_or_before(pairs, prev_cut)
        mar       = _at_or_before(pairs, mar31)
        week_pct  = ((current - prev_week) / prev_week) if (prev_week) else None
        ytd_pct   = ((current - mar) / mar) if (mar) else None
        # Day move, against the previous session's official close as Yahoo reported
        # it — not against the previous row, which holds whatever the worker last
        # wrote that day (an intra-day print, not the close) and would silently
        # compare across a gap. None where no prior close is stored: the manual GS
        # bonds and the monthly IMF/FRED series have none, and a monthly reading has
        # no "day" move to show.
        prev_close = prev_close_of.get(code)
        day_pct    = ((current - prev_close) / prev_close) if prev_close else None
        out.append({"code": code, "label": label_of[code], "unit": unit_of[code],
                    "current": current, "prev_week": prev_week, "mar31": mar,
                    "prev_close": prev_close,
                    "week_pct": week_pct, "ytd_pct": ytd_pct, "day_pct": day_pct,
                    # Date of the CURRENT reading. Daily series are always ~today, but
                    # monthly ones (IMF inflation) can be a couple of months back, and
                    # the UI has to label them rather than imply they're live.
                    "as_of": pairs[-1][0].isoformat()})
    return out


def _fy_start(as_of: date) -> date:
    """1-Apr that opens the current financial year."""
    return date(as_of.year if as_of.month >= 4 else as_of.year - 1, 4, 1)


def _avg_cost_realised(seq: list, fy_start: date) -> list:
    """
    Average-cost realised P&L from a chronological buy/sell sequence for one security.
    Each item: {date, kind: 'buy'|'sell', units, amount, name, group}.
    Records a realised row for every sell ON/AFTER fy_start.  If a sell has no known
    cost basis (no prior buys in the data — common when CAS omits old purchases),
    purchase_amount/pnl are left None rather than overstating the gain.
    """
    held, cost, out = 0.0, 0.0, []
    for t in seq:
        u   = abs(float(t["units"] or 0))
        amt = abs(float(t["amount"] or 0))
        if t["kind"] == "buy":
            held += u; cost += amt
            continue
        # sell
        if held > 1e-9 and u > 0:
            avg       = cost / held
            sold      = min(u, held)
            cost_sold = avg * sold
            held -= sold; cost -= cost_sold
        else:
            cost_sold = None
        pnl = (amt - cost_sold) if cost_sold is not None else None
        if t["date"] >= fy_start:
            ret = (pnl / cost_sold) if (pnl is not None and cost_sold) else None
            out.append({"group": t["group"], "security_name": t["name"],
                        "purchase_amount": cost_sold, "sale_date": t["date"],
                        "sale_amount": amt, "pnl": pnl, "return_pct": ret})
    return out
def _is_long_term_equity(buy_dt: date, sell_dt: date) -> bool:
    """Indian listed-equity long-term test: held for MORE than 12 months.

    The 12-month anniversary of the purchase is the boundary — a sale strictly
    after it is long-term (LTCG); on/before it is short-term (STCG). Computed on
    calendar months (not a flat 365 days) so leap years don't misclassify; a
    Feb-29 purchase rolls to Feb-28 of the anniversary year.
    """
    try:
        anniversary = buy_dt.replace(year=buy_dt.year + 1)
    except ValueError:  # Feb 29 → Feb 28 next year
        anniversary = buy_dt.replace(year=buy_dt.year + 1, day=28)
    return sell_dt > anniversary


def _fifo_realised_grouped(seq: list, fy_start: date, actions: dict | None = None) -> list:
    """
    FIFO realised P&L for ONE entity's chronological Indian-equity trade stream.

    Each sell is matched against the OLDEST open buy lots first (FIFO), mirroring
    broker-console realised P&L and Indian capital-gains reporting — so the number
    finally agrees with the client's own broker statement. This replaces the old
    average-cost method (_avg_cost_realised_grouped) for Indian equity only.

    Gross, price-to-price: gain on a matched slice = qty * (sell_price - buy_price),
    using fill prices only (brokerage/STT/other charges are NOT deducted). Each
    matched slice is split into short-term vs long-term via the lot's own buy date
    (see _is_long_term_equity), so every emitted sell carries st_pnl / lt_pnl.

    Buys before fy_start still establish lot basis; only sells on/after fy_start
    emit a row. One row per qualifying sell (aggregation into per-symbol subtotals,
    if wanted, is a display concern for the caller).

    If a sell cannot be fully covered by known lots (purchase history is
    incomplete), the whole row is flagged unknown — purchase_amount / pnl / st_pnl /
    lt_pnl are left None rather than overstating the gain, matching the
    non-overstatement policy of the average-cost engine.

    seq items: {date, kind:'buy'|'sell', units, price, name, sec}. Prices are the
    per-share INR fill price (already INR — the caller scopes to currency='INR').
    Emits: {group, security_name, purchase_amount, sale_date, sale_amount,
            pnl, st_pnl, lt_pnl, return_pct}.
    """
    lots: dict = defaultdict(deque)   # sec -> deque([[buy_date, qty, price], ...])
    out: list = []
    # Pending split/bonus events per security, consumed as the timeline reaches them.
    # Without this a 1:1 bonus makes the holder sell twice what the books show, and the
    # sell is dropped as "unknown" — the quantity is real, it just never arrived as a
    # BUY because the depository credited it. See workers/corporate_actions.py.
    pending = {s: list(v) for s, v in (actions or {}).items()}

    for t in seq:
        sec = t["sec"]
        qty = abs(float(t["units"] or 0))
        px  = float(t["price"] or 0)
        if qty <= 1e-9:
            continue

        # Must run BEFORE the trade is processed: every lot currently open was bought
        # before this date, and anything bought after the ex-date was already scaled.
        if pending.get(sec):
            _ca_apply(lots[sec], pending[sec], t["date"])

        if t["kind"] == "buy":
            lots[sec].append([t["date"], qty, px])
            continue

        # sell — consume FIFO lots for this security
        remaining       = qty
        sale_amount     = qty * px
        purchase_amount = 0.0
        st_pnl          = 0.0
        lt_pnl          = 0.0
        dq              = lots[sec]
        while remaining > 1e-9 and dq:
            lot   = dq[0]
            take  = min(remaining, lot[1])
            gain  = take * (px - lot[2])
            purchase_amount += take * lot[2]
            if _is_long_term_equity(lot[0], t["date"]):
                lt_pnl += gain
            else:
                st_pnl += gain
            lot[1]    -= take
            remaining -= take
            if lot[1] <= 1e-9:
                dq.popleft()

        if t["date"] < fy_start:
            continue

        if remaining > 1e-9:                       # not enough basis to cover the sell
            out.append({"group": t.get("group", "Equity"), "security_name": t["name"],
                        "purchase_amount": None, "sale_date": t["date"],
                        "sale_amount": sale_amount, "pnl": None,
                        "st_pnl": None, "lt_pnl": None, "return_pct": None})
        else:
            pnl = st_pnl + lt_pnl
            ret = (pnl / purchase_amount) if purchase_amount else None
            out.append({"group": t.get("group", "Equity"), "security_name": t["name"],
                        "purchase_amount": purchase_amount, "sale_date": t["date"],
                        "sale_amount": sale_amount, "pnl": pnl,
                        "st_pnl": st_pnl, "lt_pnl": lt_pnl, "return_pct": ret})
    return out


def _fx_rate_on(conn, currency: str, on_date: date) -> Optional[float]:
    """Currency→INR rate as of a date (nearest rate on/before; else earliest available).

    Trade-date FX: each leg of a foreign trade is valued at the rate on its own date,
    so realised P&L captures currency movement as well as price movement. Accuracy of
    historical legs depends on fx_rate being backfilled (see fx_backfill_worker.py);
    returns None when no rate exists at all (caller skips the leg).
    """
    if not currency or currency.upper() == "INR":
        return 1.0
    cur = conn.cursor()
    try:
        cur.execute("""SELECT rate FROM fx_rate
                       WHERE from_currency = %s AND to_currency = 'INR' AND rate_date <= %s
                       ORDER BY rate_date DESC LIMIT 1""", (currency, on_date))
        row = cur.fetchone()
        if not row:  # no rate on/before the date — fall back to the earliest we have
            cur.execute("""SELECT rate FROM fx_rate
                           WHERE from_currency = %s AND to_currency = 'INR'
                           ORDER BY rate_date ASC LIMIT 1""", (currency,))
            row = cur.fetchone()
    finally:
        cur.close()
    return float(row["rate"]) if row else None


# asset_class → realised-sheet subgroup. Driven off security_master.asset_class
# rather than security_type so the sheet buckets the same way the holdings
# sections do: gold/silver funds are precious metals, not Equity, and debt/liquid
# both land in Fixed Income (the old MF_DEBT-only test filed liquid under Equity).
# Arbitrage stays Equity — that is how it is taxed, which is what this sheet is for.
_MF_REALISED_GROUP = {
    "FIXED_INCOME": "Fixed Income",
    "GOLD_SILVER":  "Commodities",
    "ALTERNATES":   "Alternates",
}


def _mf_realised_group(asset_class: Optional[str], security_type: Optional[str]) -> str:
    ac = (asset_class or "").upper().strip()
    if ac in _MF_REALISED_GROUP:
        return _MF_REALISED_GROUP[ac]
    if ac:
        return "Equity"
    # asset_class missing (pre-classification rows) — fall back to the old test.
    return "Fixed Income" if security_type == "MF_DEBT" else "Equity"


def _fetch_realised_gains(conn, entity_ids: list, as_of: date, *,
                          since_inception: bool = False,
                          include_switches: bool = True,
                          by_broker: bool = False) -> list:
    """
    Realised gains for the given entity/entities.
    MF — auto from mf_transaction (REDEMPTION/SWITCH_OUT vs avg cost).
    Equity — from stock_transaction (SELL vs avg cost) once trades are imported.

    since_inception   — when True, count every sell on/after inception instead of
                        only the current FY (defaults to FY-to-date).
    include_switches  — when True (default), SWITCH_IN/SWITCH_OUT count as buys/sells;
                        when False they are dropped entirely (only real
                        PURCHASE/REDEMPTION flows feed cost basis and realisations).
    by_broker         — when True, FIFO/avg-cost lots are partitioned per demat
                        (broker), so a stock held at two brokers realises against
                        its own lots at each. Every emitted row carries `broker`
                        (None for MF / real estate, which have no demat). The
                        default per-entity view leaves broker None and is unchanged.
    """
    fy = date(1900, 1, 1) if since_inception else _fy_start(as_of)
    out: list = []
    cur = conn.cursor()
    ph = ",".join(["%s"] * len(entity_ids))

    # `category` is the asset bucket the Realised Gains page groups by (Equity vs
    # Mutual Funds vs …). It is intentionally separate from `group` (Fixed Income /
    # Equity / Alternates) which the XLSX realised sheet buckets by, so MF and stock
    # rows stay distinguishable on the page without disturbing the report.
    def _add(rows, category, broker=None):
        for r in rows:
            r["category"] = category
            r["broker"]   = broker
            out.append(r)

    # ---- MF ----
    # Mutual funds are held via CAS folios/AMCs, not a broker demat, so they carry
    # no broker even in the by_broker view.
    cur.execute(f"""
        SELECT t.security_id, sm.security_name, sm.security_type, sm.asset_class,
               t.transaction_date AS d, t.transaction_type AS tt, t.amount, t.units
        FROM mf_transaction t JOIN security_master sm ON sm.id = t.security_id
        WHERE t.entity_id IN ({ph})
        ORDER BY t.security_id, t.transaction_date
    """, entity_ids)
    BUY  = {"PURCHASE", "PURCHASE_SIP"} | ({"SWITCH_IN"}  if include_switches else set())
    SELL = {"REDEMPTION"}               | ({"SWITCH_OUT"} if include_switches else set())
    by_sec: dict = defaultdict(list)
    for r in cur.fetchall():
        by_sec[r["security_id"]].append(r)
    for txns in by_sec.values():
        seq = []
        for r in txns:
            kind = "buy" if r["tt"] in BUY else ("sell" if r["tt"] in SELL else None)
            if not kind:
                continue
            grp = _mf_realised_group(r["asset_class"], r["security_type"])
            seq.append({"date": r["d"], "kind": kind, "units": r["units"],
                        "amount": r["amount"], "name": r["security_name"], "group": grp})
        _add(_avg_cost_realised(seq, fy), "Mutual Funds")

    # ---- Indian equity (stock_transaction, currency='INR'; empty until imports) ----
    # FIFO lot-matching per entity as one chronological stream across all stocks:
    # each sell consumes the oldest open buy lots, so realised P&L mirrors the broker
    # console and Indian capital-gains reporting, and each sell splits into short- vs
    # long-term (see _fifo_realised_grouped). Gross, price-to-price (per-share `price`;
    # charges not deducted). Scoped to currency='INR' so Vested (USD) rows that also
    # live in stock_transaction are excluded here and handled by the foreign branch /
    # equity_trade_ledger. Ordering within a date falls back to insert order (id).
    try:
        cur.execute(f"""
            SELECT t.entity_id, t.security_id, sm.security_name, sm.isin,
                   t.transaction_date AS d, t.transaction_type AS tt,
                   t.price, t.quantity AS units,
                   -- Imported tradebook rows leave `broker` NULL and carry the executing
                   -- broker in `source` (zerodha/angel_one/dhan); only live/snapshot/manual
                   -- rows populate `broker` directly. Derive it the same way the rest of the
                   -- codebase does (fy_returns_worker, equity_txn_metrics_worker) so the demat
                   -- view isn't dominated by a "No demat" bucket of every imported trade.
                   COALESCE(t.broker,
                            CASE WHEN t.source IN ('zerodha','angel_one','dhan')
                                 THEN t.source END) AS broker
            FROM stock_transaction t JOIN security_master sm ON sm.id = t.security_id
            WHERE t.entity_id IN ({ph})
              AND COALESCE(t.currency, 'INR') = 'INR'
            ORDER BY t.entity_id, t.transaction_date, t.id
        """, entity_ids)
        srows = cur.fetchall()
    except Exception:
        conn.rollback(); srows = []
    try:
        ca_actions = _ca_load(cur)
    except Exception:
        conn.rollback(); ca_actions = {}
    # Gold/silver ETFs and SGBs live in stock_transaction like any other scrip —
    # security_master carries no asset_class for them (equity_holding does, and a
    # fully-sold position is pruned from there), so classify from symbol/ISIN with
    # the shared classifier, admin overrides included.
    try:
        _ac_ovr = _ac_overrides(cur)
    except Exception:
        conn.rollback(); _ac_ovr = {}
    # Per-entity by default; per (entity, broker) in the demat view so a stock held
    # at two brokers FIFO-matches against its own lots at each. The broker is not part
    # of the FIFO lot key itself — partitioning the input stream already keeps the
    # brokers' lots apart.
    eq_by_entity: dict = defaultdict(list)
    for r in srows:
        key = (r["entity_id"], r["broker"]) if by_broker else (r["entity_id"], None)
        eq_by_entity[key].append(r)
    for (_eid, bk), txns in eq_by_entity.items():
        seq = []
        for r in txns:
            tt = (r["tt"] or "").upper()
            kind = "buy" if tt in ("BUY", "B", "PURCHASE") else ("sell" if tt in ("SELL", "S", "SALE") else None)
            if not kind:
                continue
            cls = _ac_classify(r["security_name"], r["isin"], _ac_ovr)
            seq.append({"date": r["d"], "kind": kind, "units": r["units"],
                        "price": r["price"], "name": r["security_name"],
                        "group": "Commodities" if cls in _AC_COMMODITY else "Equity",
                        "sec": r["security_id"]})
        _add(_fifo_realised_grouped(seq, fy, ca_actions), "Equity", broker=bk)

    # ---- Foreign equity (equity_trade_ledger; native cash flows → INR at trade-date FX) ----
    # Switches do not exist for brokers, so include_switches is irrelevant here.
    # cash_flow_native is signed (BUY negative / SELL positive); we avg-cost on |amount|
    # converted to INR at each leg's own trade-date rate, so realised P&L embeds FX gain.
    try:
        cur.execute(f"""
            SELECT symbol, isin, trade_date AS d, side,
                   quantity AS units, currency, cash_flow_native, broker
            FROM equity_trade_ledger
            WHERE entity_id IN ({ph})
            ORDER BY symbol, trade_date
        """, entity_ids)
        frows = cur.fetchall()
    except Exception:
        conn.rollback(); frows = []
    fe_by_sec: dict = defaultdict(list)
    for r in frows:
        key = (r["symbol"], r["isin"], r["broker"]) if by_broker else (r["symbol"], r["isin"], None)
        fe_by_sec[key].append(r)
    for (_sym, _isin, bk), txns in fe_by_sec.items():
        seq = []
        for r in txns:
            side = (r["side"] or "").upper()
            kind = "buy" if side == "BUY" else ("sell" if side == "SELL" else None)
            if not kind:
                continue
            fx = _fx_rate_on(conn, r["currency"], r["d"])
            if fx is None:
                continue  # no rate available — cannot express this leg in INR
            amt_inr = abs(float(r["cash_flow_native"] or 0)) * fx
            seq.append({"date": r["d"], "kind": kind, "units": r["units"],
                        "amount": amt_inr, "name": r["symbol"], "group": "Foreign Equity"})
        _add(_avg_cost_realised(seq, fy), "Foreign Equity", broker=bk)

    # ---- PMS (pms_realised; realised P&L already computed per lot) ----
    # No avg-cost here. Nuvama/Zerodha publish a realised capital-gains statement
    # and those rows are imported as-is; ICICI publishes no such statement, so
    # icici_pms_worker FIFO-matches its transaction history instead (only when
    # that history reconciles against current holdings). Either way the cost,
    # proceeds and P&L arrive per lot. INR only.
    try:
        cur.execute(f"""
            SELECT security_name, sale_date AS d, purchase_amount, sale_amount, pnl
            FROM pms_realised
            WHERE entity_id IN ({ph}) AND sale_date >= %s
            ORDER BY sale_date
        """, entity_ids + [fy])
        prows = cur.fetchall()
    except Exception:
        conn.rollback(); prows = []
    for r in prows:
        pnl = float(r["pnl"]) if r["pnl"] is not None else None
        pa  = float(r["purchase_amount"]) if r["purchase_amount"] is not None else None
        ret = (pnl / pa) if (pnl is not None and pa) else None
        out.append({"group": "PMS", "category": "PMS", "security_name": r["security_name"],
                    "purchase_amount": pa, "sale_date": r["d"],
                    "sale_amount": float(r["sale_amount"]) if r["sale_amount"] is not None else None,
                    "pnl": pnl, "return_pct": ret})

    cur.close()
    return out


# ── register workbook (properties + art / collectibles) ──────────────────────
#
# These two registers are deliberately kept OUT of the portfolio reports and out
# of the dashboard totals (owner decision: they are standalone asset registers,
# not part of the tradeable book). They used to have sections in the entity and
# master workbooks that silently rendered empty — properties moved to their own
# `property` table in the 2026-07-13 rebuild and the report kept reading the old
# manual_input category, while the Art line read `art` after everything had been
# recategorised to `collectibles`. Both now live here, read from the right source,
# and the dead sections are gone from the portfolio reports.
#
# Every sheet gets an Excel AutoFilter on its header row, so each column can be
# filtered/sorted from its own dropdown.

# Valuation must match what the Properties page shows, or the register is worse
# than useless. Both constants are imported from the modules that own them rather
# than re-declared here — a second copy of 1.75 is exactly how the two drift apart.
try:                                          # worker runs with mis-portal/ on the path
    from property_docs import FAIR_VALUE_MULTIPLIER
except ImportError:                           # pragma: no cover — layout fallback
    from mis_portal.property_docs import FAIR_VALUE_MULTIPLIER  # type: ignore

OLD_LEASE_OWNER_SHARE = 0.5   # statutory sitting tenant holds the other half


def _fetch_property_register(conn):
    """Every property with its holder, valuation inputs and effective fair value.

    The value expression mirrors _fetch_property_overview_rows in main.py exactly:
    sale_price once sold; else land (hand-entered market_land_value, else
    area x RRR x the fair-value multiplier) plus summed floor costings; halved for
    an old statutory lease. Getting any of those wrong would put a number in the
    register that the Properties page contradicts.
    """
    val_expr = (
        f"CASE WHEN p.sold THEN p.sale_price ELSE "
        f"(COALESCE(p.market_land_value, p.area * p.rrr * {FAIR_VALUE_MULTIPLIER}) "
        f" + COALESCE(fv.bval, 0)) "
        f"* CASE WHEN p.is_old_lease THEN {OLD_LEASE_OWNER_SHARE} ELSE 1 END END"
    )
    cur = conn.cursor()
    cur.execute(f"""
        WITH fv AS (
            SELECT property_id,
                   SUM(COALESCE(built_up_area, area) * rate_per_unit) AS bval
            FROM property_floor WHERE rate_per_unit IS NOT NULL
            GROUP BY property_id
        )
        SELECT p.name, p.property_type, e.name AS holder_name, p.ownership,
               p.village, p.taluka, p.address, p.survey_no, p.property_no,
               p.area, p.area_unit, p.built_up_area, p.tenure, p.is_old_lease,
               p.acquisition_date, p.purchase_price, p.rrr, p.market_land_value,
               p.sold, p.sale_date, p.sale_price, p.notes,
               {val_expr} AS fair_value
        FROM   property p
        JOIN   property_entity e ON e.id = p.holder_id
        LEFT   JOIN fv ON fv.property_id = p.id
        ORDER  BY e.sort_order, e.name, p.sold, p.name
    """)
    rows = cur.fetchall()
    cur.close()
    return rows


def _fetch_art_register(conn):
    """Latest art + collectibles entry per (entity, label), with painter/location
    detail. Both categories together: the split is a page-level distinction, and
    for a register you want the one list."""
    cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT ON (m.entity_id, m.category, m.label)
               m.category, m.label, e.entity_name, m.cost, m.current_value,
               m.currency, m.inception_date, m.notes, m.updated_at,
               d.painter_name, d.location, d.seller_name, d.seller_address
        FROM   manual_input m
        JOIN   entity e ON e.id = m.entity_id
        LEFT   JOIN art_detail d
               ON d.entity_id = m.entity_id AND d.label = m.label
        WHERE  m.category IN ('art', 'collectibles')
        ORDER  BY m.entity_id, m.category, m.label, m.updated_at DESC
    """)
    rows = cur.fetchall()
    cur.close()
    return rows


def _register_sheet(ws, title: str, headers: list, rows: list, widths: list):
    """Title row, filterable header row, then the data. Shared by both registers
    so they read as one document."""
    ncols = len(headers)
    # Same gold vocabulary as the portfolio sheets, so the registers read as part of
    # the same document rather than a stray export.
    _rf.paint_row(ws, 1, 1, ncols, font=_rf.F_TITLE16, fill=_rf.GOLD_PALE,
                  border=_rf.B_BAND, align=_rf.AL_CENTER)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1).value = title
    ws.row_dimensions[1].height = 24

    _rf.paint_row(ws, 2, 1, ncols, font=_rf.F_HDR_DARK, fill=_rf.AMBER,
                  border=_rf.B_HDR, align=_rf.AL_CENTER_WRAP)
    for i, h in enumerate(headers, start=1):
        ws.cell(row=2, column=i).value = h
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[2].height = 30

    for r, values in enumerate(rows, start=3):
        _rf.paint_row(ws, r, 1, ncols, font=_rf.F_BODY, border=_rf.B_THIN)
        for i, v in enumerate(values, start=1):
            c = ws.cell(row=r, column=i, value=v)
            if isinstance(v, (int, float)):
                c.number_format = _rf.MONEY
                c.alignment = _rf.AL_RIGHT
            else:
                c.alignment = _rf.AL_LEFT_WRAP

    # The filter dropdowns the registers are meant to be worked through.
    last = f"{get_column_letter(ncols)}{max(2, len(rows) + 2)}"
    ws.auto_filter.ref = f"A2:{last}"
    ws.freeze_panes = "A3"
    return ws


def build_register_workbook(conn, as_of: date):
    """Standalone register workbook: Properties + Art & Collectibles, all entities.
    Returns None when both registers are empty, so no pointless file is written."""
    props = _fetch_property_register(conn)
    art   = _fetch_art_register(conn)
    if not props and not art:
        return None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if props:
        headers = ["Property", "Type", "Holder", "Ownership", "Village", "Taluka",
                   "Address", "Survey No", "Property No", "Area", "Unit",
                   "Built-up Area", "Tenure", "Old Lease", "Acquired",
                   "Purchase Price (₹)", "RRR (₹)", "Fair Value (₹)", "Status",
                   "Sale Date", "Sale Price (₹)", "Notes"]
        widths  = [28, 12, 20, 14, 14, 14, 30, 14, 14, 10, 8, 12, 12, 10, 12,
                   16, 14, 16, 10, 12, 16, 30]
        rows = []
        for p in props:
            rows.append([
                p["name"], p["property_type"], p["holder_name"], p["ownership"],
                p["village"], p["taluka"], p["address"], p["survey_no"],
                p["property_no"],
                float(p["area"]) if p["area"] is not None else None,
                p["area_unit"],
                float(p["built_up_area"]) if p["built_up_area"] is not None else None,
                p["tenure"],
                "Yes" if p["is_old_lease"] else "",
                p["acquisition_date"].isoformat() if p["acquisition_date"] else None,
                float(p["purchase_price"]) if p["purchase_price"] is not None else None,
                float(p["rrr"]) if p["rrr"] is not None else None,
                float(p["fair_value"]) if p["fair_value"] is not None else None,
                "Sold" if p["sold"] else "Held",
                p["sale_date"].isoformat() if p["sale_date"] else None,
                float(p["sale_price"]) if p["sale_price"] is not None else None,
                p["notes"],
            ])
        _register_sheet(wb.create_sheet("Properties"),
                        f"PROPERTY REGISTER — as of {as_of:%d %b %Y}",
                        headers, rows, widths)

    if art:
        headers = ["Item", "Class", "Entity", "Painter", "Kept At", "Acquired",
                   "Purchase Price", "Current Valuation", "Currency",
                   "Seller", "Seller Address", "Last Updated", "Notes"]
        widths  = [30, 14, 20, 22, 20, 12, 16, 18, 10, 22, 30, 14, 30]
        rows = []
        for a in art:
            rows.append([
                a["label"],
                "Collectible" if a["category"] == "collectibles" else "Art",
                a["entity_name"], a["painter_name"], a["location"],
                a["inception_date"].isoformat() if a["inception_date"] else None,
                float(a["cost"]) if a["cost"] is not None else None,
                float(a["current_value"]) if a["current_value"] is not None else None,
                a["currency"], a["seller_name"], a["seller_address"],
                a["updated_at"].strftime("%Y-%m-%d") if a["updated_at"] else None,
                a["notes"],
            ])
        _register_sheet(wb.create_sheet("Art & Collectibles"),
                        f"ART & COLLECTIBLES REGISTER — as of {as_of:%d %b %Y}",
                        headers, rows, widths)

    return wb


def _entity_pages(conn) -> list[dict]:
    """
    Every page the client wants a Weekly + Realised sheet for, in sheet order:
    the multi-entity PAN groups first, then each entity on its own.

    Returns [{"label", "entity_ids", "entity_id"}] — entity_id is None for a group.
    """
    # The client names three groups on the reports — Dhruv, Harsh and Stuti. Those get
    # a group page; the remaining PAN groups hold a single entity, whose group page
    # would be a byte-for-byte duplicate of its own entity page, so they are skipped.
    pages = [{"label": g["label"], "entity_ids": g["entity_ids"], "entity_id": None}
             for g in _all_pan_groups(conn) if g["label"].endswith(" GROUP")]
    for e in _fetch_entities(conn):
        pages.append({"label": e["entity_name"], "entity_ids": [e["id"]],
                      "entity_id": e["id"]})
    return pages


def _page_payload(conn, entity_ids: list, as_of: date) -> dict:
    """Everything one Weekly/Realised page pair needs.

    Realised gains are fetched TWICE on purpose. The weekly page's asset-class lines
    use the entity-level view, so they agree with the Realised Gains page in the
    portal. The realised DETAIL page uses the by-demat view, because the client's
    sheet lists direct-equity realisations under the broker they happened at, which
    needs FIFO lots partitioned per demat. The two totals differ by around 1% — a
    stock held at two brokers realises against each demat's own lots — and that
    difference is expected, not a discrepancy to reconcile away.
    """
    return {
        "bundle":       _bundle_for(conn, entity_ids, as_of),
        "realised":     _fetch_realised_gains(conn, entity_ids, as_of),
        "realised_dtl": _fetch_realised_gains(conn, entity_ids, as_of, by_broker=True),
        "dividends":    _fetch_dividends(conn, entity_ids, as_of),
    }


def _realised_by_class(realised: list) -> dict:
    """Realised P&L totalled into the client's three asset-class lines."""
    out = {"fixed": 0.0, "equity": 0.0, "alt": 0.0}
    for r in realised or []:
        g = (r.get("group") or "").lower()
        pnl = r.get("pnl")
        pnl = float(pnl) if pnl is not None else 0.0
        if g in ("fixed income", "mutual funds"):
            out["fixed"] += pnl
        elif g in ("commodities", "alternates"):
            out["alt"] += pnl
        else:
            out["equity"] += pnl
    return out


# Excel caps sheet names at 31 characters and forbids : \ / ? * [ ].
_SHEET_BAD = re.compile(r"[:\\/?*\[\]]")


def _sheet_name(label: str, suffix: str) -> str:
    """'<label> <suffix>', trimmed to Excel's 31-char limit from the LABEL end so the
    suffix ('Weekly Report' / 'Realised P&L') stays readable."""
    name = _SHEET_BAD.sub("-", f"{label} {suffix}")
    if len(name) <= 31:
        return name
    keep = 31 - len(suffix) - 1
    return f"{_SHEET_BAD.sub('-', label)[:keep].strip()} {suffix}"


def build_master_workbook(conn, as_of: date):
    """
    The consolidated MIS workbook, in the client's own tab order:

      1. Equity Daily Print        — group total + per-entity, domestic then foreign
      2. All Assets Daily MIS      — the whole book by asset class
      3. All Entities Weekly Report— one column block per PAN group + ALL RAJANI
      4. <Label> Weekly Report     — one per PAN group, then one per entity
      5. <Label> Realised P&L      — the same list again

    Every sheet is drawn by workers/report_sheets.py, whose layout is taken cell for
    cell off the client's master format workbook.
    """
    wb = openpyxl.Workbook()
    entities   = _fetch_entities(conn)
    benchmarks = _fetch_benchmarks(conn, as_of)
    all_ids    = [e["id"] for e in entities]

    # 1) Equity Daily Print
    ws = wb.active
    ws.title = "Equity Daily Print"
    dom, forn = _edp_holdings(conn, {e["entity_name"]: e["id"] for e in entities})
    _render_equity_daily_print(ws, as_of, "Rajani Group", dom, forn)

    # 2) All Assets Daily MIS — one bundle covering every entity.
    combined = _bundle_for(conn, all_ids, as_of)
    build_all_assets_daily(wb.create_sheet("All Assets Daily MIS"), as_of,
                           classify_bundle(combined), benchmarks)

    # 3) All Entities Weekly Report — a column block per PAN group (singletons
    #    included, so the ALL RAJANI total agrees with the per-entity pages).
    group_blocks = []
    for g in _all_pan_groups(conn):
        realised = _fetch_realised_gains(conn, g["entity_ids"], as_of)
        group_blocks.append({
            "label":    g["label"],
            "sec":      classify_bundle(_bundle_for(conn, g["entity_ids"], as_of)),
            "realised": _realised_by_class(realised),
        })
    build_all_entities_weekly(wb.create_sheet("All Entities Weekly Report"),
                              as_of, group_blocks, benchmarks)

    # 4+5) Weekly + Realised page per group, then per entity.
    for page in _entity_pages(conn):
        d = _page_payload(conn, page["entity_ids"], as_of)
        build_weekly_report(wb.create_sheet(_sheet_name(page["label"], "Weekly Report")),
                            page["label"], d["bundle"], as_of,
                            benchmarks=benchmarks, realised=d["realised"],
                            staleness=_staleness_note(conn, as_of, page["entity_id"]))
        build_realised_pnl(wb.create_sheet(_sheet_name(page["label"], "Realised P&L")),
                           page["label"], as_of,
                           realised=d["realised_dtl"], dividends=d["dividends"])
    return wb


def build_entity_workbook(conn, entity: dict, as_of: date):
    """One standalone workbook for a single entity: its Weekly Report page and its
    FY Realised P&L page, both in the client's format. Saved as ENTITYNAME-DATE.xlsx."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    label = entity["entity_name"]
    d = _page_payload(conn, [entity["id"]], as_of)
    build_weekly_report(wb.create_sheet("Weekly Report"), label, d["bundle"], as_of,
                        benchmarks=_fetch_benchmarks(conn, as_of),
                        realised=d["realised"],
                        staleness=_staleness_note(conn, as_of, entity["id"]))
    build_realised_pnl(wb.create_sheet("Realised P&L"), label, as_of,
                       realised=d["realised_dtl"], dividends=d["dividends"])
    return wb


def _safe_filename(name: str) -> str:
    """Filesystem-safe entity name for report filenames (spaces/slashes → '_')."""
    return re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("_") or "entity"


# ── public entry point ────────────────────────────────────────────────────────

def generate_reports(conn, generated_by_user_id: Optional[int] = None) -> list[dict]:
    """
    Generate (1) the consolidated master MIS workbook (all entities) and (2) one
    standalone workbook per entity, named ENTITYNAME-DATE.xlsx. Realised P&L is
    embedded at the bottom of each Weekly sheet — never a separate sheet/report.
    Registers every file in generated_report and returns a row per file.
    """
    as_of  = date.today()
    folder = os.path.join(REPORTS_DIR, as_of.strftime("%Y-%m-%d"))
    os.makedirs(folder, exist_ok=True)

    cur = conn.cursor()
    results: list[dict] = []

    def _register(report_type, entity_id, entity_name, fname, fpath):
        cur.execute("""
            INSERT INTO generated_report
                (report_type, entity_id, entity_name, filename, filepath, as_of_date, generated_by, generated_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, NOW())
            RETURNING id
        """, (report_type, entity_id, entity_name, fname, fpath, as_of, generated_by_user_id))
        results.append({"id": cur.fetchone()["id"], "type": report_type,
                        "entity": entity_name, "filename": fname, "path": fpath})

    # 1) Consolidated master workbook (all entities)
    m_name = f"MIS-Report_{as_of.strftime('%Y%m%d')}.xlsx"
    m_path = os.path.join(folder, m_name)
    build_master_workbook(conn, as_of).save(m_path)
    _register("master", None, "MIS Report — All Entities", m_name, m_path)

    # 2) One standalone workbook per entity (ENTITYNAME-DATE.xlsx)
    for e in _fetch_entities(conn):
        ename = e["entity_name"]
        fname = f"{_safe_filename(ename)}-{as_of.strftime('%Y-%m-%d')}.xlsx"
        fpath = os.path.join(folder, fname)
        build_entity_workbook(conn, e, as_of).save(fpath)
        _register("individual", e["id"], ename, fname, fpath)

    # 3) Register workbook — properties + art/collectibles, all entities. These are
    # standalone registers kept out of the portfolio reports and totals, so they get
    # their own file rather than an empty section in everyone else's.
    reg_wb = build_register_workbook(conn, as_of)
    if reg_wb is not None:
        r_name = f"Registers_{as_of.strftime('%Y%m%d')}.xlsx"
        r_path = os.path.join(folder, r_name)
        reg_wb.save(r_path)
        _register("registers", None, "Registers — Properties & Art", r_name, r_path)

    conn.commit()
    cur.close()
    return results


if __name__ == "__main__":
    from dotenv import load_dotenv
    load_dotenv("/var/www/mis-portal/.env")
    conn = psycopg2.connect(
        host=os.getenv("DB_HOST", "localhost"),
        database=os.getenv("DB_NAME", "mis_portal"),
        user=os.getenv("DB_USER", "postgres"),
        password=os.getenv("DB_PASSWORD", ""),
        cursor_factory=RealDictCursor,
    )
    results = generate_reports(conn)
    for r in results:
        print(f"✅  {r['type']:8s} {r['entity']:30s} → {r['filename']}")
    conn.close()
