#!/usr/bin/env python3
"""
Sheet builders that reproduce the client's MIS workbook layout exactly.

Every block position, header string, merge, fill and number format below was read
off the client's master format file (MIS-CURRENT-FORMAT.xlsx) — see report_format.py
for the style vocabulary. The row CONTENT is live: each section lists whatever the
entity actually holds today, in the template's own blocks and columns.

Five sheet shapes, matching the client's five tabs:

  equity_daily_print   — 7+ blocks of 13 columns (A-M), domestic then foreign
  all_assets_daily_mis — benchmarks, then Fixed Income / Equity / Alternates (A-H)
  all_entities_weekly  — 5 columns per PAN group + an ALL RAJANI GROUP total
  weekly_report        — the "Dhruv Weekly Report" page, cloned per entity (A-P)
  realised_pnl         — the "FY2627 Realised Profit & Loss" page, per entity (A-G)
"""
from datetime import date
from collections import defaultdict

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from workers.report_format import (
    GOLD_DARK, GOLD_LIGHT, AMBER, GOLD_PALE, CREAM, PALE_BLUE, NO_FILL,
    F_TITLE, F_TITLE16, F_BANNER, F_BANNER20, F_HDR, F_HDR12, F_HDR7, F_HDR_DARK,
    F_SUBTOTAL, F_GRANDTOT, F_SECTION, F_BODY, F_BODY12, F_BODY_BOLD,
    F_DATE_SM, F_NOTE, F_REMARK,
    B_THIN, B_BOX, B_BAND, B_TOTAL, B_HDR, B_NONE,
    MONEY, MONEY_TOT, PCT2, PCT0, QTY, PRICE, DATE_LONG, DATE_SHORT, DATE_DMY,
    AL_CENTER_WRAP, AL_CENTER, AL_LEFT, AL_LEFT_WRAP, AL_RIGHT,
    style_cell, paint_row, banner, set_widths, note,
)

# ── small numeric helpers ─────────────────────────────────────────────────────

def _f(row, key):
    """Float value of `key`, or None. Treats a missing key and a NULL alike."""
    v = row.get(key) if isinstance(row, dict) else None
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _sum(rows, key):
    """Sum of `key` across rows, or None when not one row carries it.

    None rather than 0 matters: a blank cell says "we hold nothing that reports
    this", while a 0 asserts the value is genuinely zero. The client's sheets lean
    on that distinction, so totals must not manufacture zeros.
    """
    vals = [_f(r, key) for r in rows]
    vals = [v for v in vals if v is not None]
    return sum(vals) if vals else None


def _pct(row, key):
    """Percent columns are stored as 0-100 in the DB but written as Excel fractions."""
    v = _f(row, key)
    return v / 100.0 if v is not None else None


def _label_of(h):
    return h.get("security_name") or h.get("symbol") or h.get("label") or ""


def _remark_of(h):
    return h.get("broker") or h.get("remarks") or h.get("notes") or ""


def _inception_of(h):
    return h.get("first_invested_date") or h.get("inception_date")


def _weekly_change(h):
    wk = _f(h, "weekly_change")
    if wk is not None:
        return wk
    cmv, prev = _f(h, "current_value"), _f(h, "prev_week_value")
    return (cmv - prev) if (cmv is not None and prev is not None) else None


def _weekly_agg(rows):
    """(weekly change, the previous-week base it was measured against) for a group.

    Summing every CMV and every previous-week value separately and subtracting is
    wrong whenever a holding has one but not the other: a PPF row with a current
    value and no prior week would book its ENTIRE balance as a week's move. So the
    change is summed per row, and the percentage base counts only the rows that
    actually contributed to it.
    """
    total, base = None, None
    for r in rows:
        wk = _weekly_change(r)
        if wk is None:
            continue
        prev = _f(r, "prev_week_value")
        total = (total or 0.0) + wk
        if prev is not None:
            base = (base or 0.0) + prev
    return total, base


def _fy_label(as_of: date) -> str:
    """'26/27' for the financial year containing as_of (Apr-Mar)."""
    y = as_of.year if as_of.month >= 4 else as_of.year - 1
    return f"{y % 100:02d}/{(y + 1) % 100:02d}"


def _mar31(as_of: date) -> date:
    """The 31-Mar that closed the previous financial year."""
    return date(as_of.year if as_of.month >= 4 else as_of.year - 1, 3, 31)


# ── fund classification ───────────────────────────────────────────────────────
# The client's sheets split MF_DEBT into Liquid / Debt / Arbitrage and MF_EQUITY
# into Index / Market / Sector-Thematic / Hybrid. The DB carries no such split —
# security_master only knows MF_DEBT, MF_EQUITY, MF_HYBRID — so the fund name is
# the signal, exactly as the client groups them by hand.

_LIQUID_HINTS    = ("liquid", "overnight", "money market")
_ARBITRAGE_HINTS = ("arbitrage", "equity saving", "eqsaving", "eq saving")
_INDEX_HINTS     = ("index", "nifty", "sensex", "s&p", "etf")
_HYBRID_HINTS    = ("asset allo", "multi asset", "children", "balanced",
                    "hybrid", "advantage", "dynamic asset")
_SECTOR_HINTS    = ("business cycle", "comma", "pharma", "technology", "infra",
                    "banking and psu", "thematic", "sector", "consumption",
                    "manufactur", "energy", "psu equity")


def _has(name: str, hints) -> bool:
    """Hint match on a fund name, punctuation-insensitive.

    AMCs spell the same fund "Multi-Asset", "Multi Asset" and "MultiAsset" across
    statements, so a raw substring test silently mis-buckets it — that is how
    "ICICI Prudential Multi-Asset Fund" ends up under Market Equity instead of
    Hybrid. Hyphens, dots, slashes and ampersands collapse to spaces first.
    """
    n = (name or "").lower()
    for ch in "-_./&()":
        n = n.replace(ch, " ")
    n = " ".join(n.split())
    return any(h in n for h in hints)


def fixed_income_bucket(h) -> str:
    """'liquid' | 'arbitrage' | 'debt' for a fixed-income MF holding."""
    name = _label_of(h)
    if _has(name, _ARBITRAGE_HINTS):
        return "arbitrage"
    if _has(name, _LIQUID_HINTS):
        return "liquid"
    return "debt"


def equity_mf_bucket(h) -> str:
    """'index' | 'market' | 'sector' | 'hybrid' for an equity/hybrid MF holding."""
    name = _label_of(h)
    if h.get("security_type") == "MF_HYBRID" or _has(name, _HYBRID_HINTS):
        return "hybrid"
    if _has(name, _INDEX_HINTS):
        return "index"
    if _has(name, _SECTOR_HINTS):
        return "sector"
    return "market"


# ── bundle → the client's own asset-class sections ────────────────────────────

def classify_bundle(bundle: dict) -> dict:
    """
    Split one _bundle_for() result into the exact section buckets the client's
    sheets use. Every key is always present (possibly an empty list) so callers can
    render a section header without first testing for it.
    """
    mf   = bundle.get("mf", [])
    man  = bundle.get("manual_by_cat", {})

    debt_mf = [h for h in mf if h.get("security_type") == "MF_DEBT"]
    eq_mf   = [h for h in mf if h.get("security_type") in ("MF_EQUITY", "MF_HYBRID")]

    # An "equity savings"/arbitrage fund is tagged MF_HYBRID or MF_EQUITY in the DB
    # but the client reports it under Fixed Income, so it is pulled across here.
    arbitrage = ([h for h in debt_mf if fixed_income_bucket(h) == "arbitrage"] +
                 [h for h in eq_mf   if _has(_label_of(h), _ARBITRAGE_HINTS)])
    eq_mf     = [h for h in eq_mf if not _has(_label_of(h), _ARBITRAGE_HINTS)]

    return {
        # ── Fixed income ──
        "liquid":    [h for h in debt_mf if fixed_income_bucket(h) == "liquid"],
        "debt":      [h for h in debt_mf if fixed_income_bucket(h) == "debt"],
        "arbitrage": arbitrage,
        "ppf":       man.get("ppf", []),

        # ── Equity ──
        "mf_index":  [h for h in eq_mf if equity_mf_bucket(h) == "index"],
        "mf_market": [h for h in eq_mf if equity_mf_bucket(h) == "market"],
        "mf_sector": [h for h in eq_mf if equity_mf_bucket(h) == "sector"],
        "mf_hybrid": [h for h in eq_mf if equity_mf_bucket(h) == "hybrid"],
        "pms":       bundle.get("pms", []) + man.get("pms", []),
        "direct":    bundle.get("eq", []) + man.get("direct_equity", []),
        "foreign":   (bundle.get("fe", []) + man.get("overseas_equity", [])
                      + man.get("overseas_fund", [])),
        "aif":       man.get("aif", []) + man.get("fno", []),

        # ── Alternates ──
        "unlisted":  man.get("unlisted", []) + man.get("startup", []),
        "forex":     man.get("forex", []) + man.get("nre_bank", []),
        "gold":      bundle.get("comm", []) + man.get("gold_etf", []),

        # ── Below the line ──
        "transit":   man.get("funds_transit", []),
        "broker":    bundle.get("cash", []) + man.get("broker_balance", []),
        "bank":      bundle.get("bank", []) + man.get("bank", []),
    }


# Which buckets roll into each of the client's lettered totals.
FIXED_BUCKETS  = ["liquid", "debt", "arbitrage", "ppf"]
EQUITY_BUCKETS = ["mf_index", "mf_market", "mf_sector", "mf_hybrid",
                  "pms", "direct", "foreign", "aif"]
ALT_BUCKETS    = ["unlisted", "forex", "gold"]
BELOW_BUCKETS  = ["transit", "broker", "bank"]


def _rows_of(sec: dict, buckets) -> list:
    out = []
    for b in buckets:
        out.extend(sec.get(b, []))
    return out


def section_totals(sec: dict) -> dict:
    """CMV / cost / prev-week / P&L aggregates per lettered total, for the summary
    block at the top of the weekly page and for the all-entities monitor."""
    def agg(rows):
        cmv  = _sum(rows, "current_value")
        prev = _sum(rows, "prev_week_value")
        cost = _sum(rows, "cost")
        pnl_ytd = _sum(rows, "pnl_ytd")
        pnl_inc = _sum(rows, "pnl_inception")
        wk, wk_base = _weekly_agg(rows)
        return {
            "rows": rows, "cmv": cmv, "prev": prev, "cost": cost,
            "mar31": _sum(rows, "market_value_as_on"),
            "wk": wk,
            "pnl_ytd": pnl_ytd, "pnl_inception": pnl_inc,
            "wk_pct": (wk / wk_base) if (wk is not None and wk_base) else None,
            "ret_ytd": (pnl_ytd / (cmv - pnl_ytd)) if (pnl_ytd is not None and cmv
                        and (cmv - pnl_ytd)) else None,
            "ret_inception": (pnl_inc / cost) if (pnl_inc is not None and cost) else None,
        }

    out = {
        "fixed":  agg(_rows_of(sec, FIXED_BUCKETS)),
        "equity": agg(_rows_of(sec, EQUITY_BUCKETS)),
        "alt":    agg(_rows_of(sec, ALT_BUCKETS)),
        "transit": agg(sec.get("transit", [])),
        "broker":  agg(sec.get("broker", [])),
        "bank":    agg(sec.get("bank", [])),
    }
    out["sub"]   = agg(out["fixed"]["rows"] + out["equity"]["rows"])
    out["grand"] = agg(out["sub"]["rows"] + out["alt"]["rows"]
                       + _rows_of(sec, BELOW_BUCKETS))
    return out


# ══════════════════════════════════════════════════════════════════════════════
#  Weekly Report  —  the client's "Dhruv Weekly Report" page, cloned per entity
# ══════════════════════════════════════════════════════════════════════════════

# Columns A-P. H is the client's 0.43-wide spacer rule between the value columns
# and the performance columns — it carries no data and must stay.
WK_NAME, WK_INCEPT, WK_PREV, WK_MAR31, WK_COST, WK_EXP, WK_CMV = 1, 2, 3, 4, 5, 6, 7
WK_GAP = 8
WK_WKCHG, WK_PNL_YTD, WK_PNL_INC = 9, 10, 11
WK_WKPCT, WK_RET_YTD, WK_RET_INC, WK_CAGR, WK_REMARK = 12, 13, 14, 15, 16
WK_LAST = 16

# The Performance Summary block at the top of the page has no inception columns, so
# the client shifts its two percentage columns one to the left of where the detail
# blocks put them: weekly-change % in K and YTD returns in L, not L and M.
WKS_WKPCT, WKS_RET_YTD = 11, 12

WK_WIDTHS = {
    "A": 39.14, "B": 11.14, "C": 14.57, "D": 14.43, "E": 16.71, "F": 12.29,
    "G": 16.43, "H": 0.43,  "I": 14.00, "J": 15.14, "K": 15.14, "L": 10.29,
    "M": 10.71, "N": 10.71, "O": 12.29, "P": 22.29,
}

# Money columns and percent columns of the detail table, for number formatting.
WK_MONEY = (WK_PREV, WK_MAR31, WK_COST, WK_CMV, WK_WKCHG, WK_PNL_YTD, WK_PNL_INC)
WK_PCT   = (WK_EXP, WK_WKPCT, WK_RET_YTD, WK_RET_INC, WK_CAGR)

# Benchmark rows of the Market Statistics block, in the client's order. The codes
# are market_benchmark.code; the labels are the client's own wording.
BENCHMARK_ROWS = [
    ("SENSEX",       "BSE - Sensex"),
    ("NIFTY",        "NSE - Nifty"),
    ("GS2032_YTM",   "7.26% GS 2032 (Benchmark) (YTM)"),
    ("GS2032_PRICE", "7.26% GS 2032 (Benchmark) (Price)"),
    ("GS2030_YTM",   "5.77% NI GS 2030 (Benchmark) (YTM)"),
    ("GS2030_PRICE", "5.77% NI GS 2030 (Benchmark) (Price)"),
]


class _WeeklyWriter:
    """Row cursor + the four cell shapes the client's weekly page is built from."""

    def __init__(self, ws, as_of, total_cmv):
        self.ws = ws
        self.row = 1
        self.as_of = as_of
        self.total_cmv = total_cmv or 0.0

    # -- generic ------------------------------------------------------------
    def blank(self, n=1):
        self.row += n

    def title(self, text, span=WK_LAST):
        """Big 22pt centred page/section title, e.g. '- FIXED INCOME INVESTMENT'."""
        r = self.row
        paint_row(self.ws, r, 1, span, font=F_TITLE, align=AL_CENTER, fill=NO_FILL)
        self.ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
        self.ws.cell(row=r, column=1).value = text
        self.ws.row_dimensions[r].height = 28
        self.row += 1
        return r

    def entity_banner(self, label):
        """The gold '<LABEL> GROUP' bar that names whose numbers follow."""
        r = self.row
        paint_row(self.ws, r, 1, WK_LAST, font=F_BANNER, fill=GOLD_DARK,
                  border=B_BAND, align=AL_CENTER)
        self.ws.merge_cells(start_row=r, start_column=WK_COST,
                            end_row=r, end_column=WK_LAST)
        self.ws.cell(row=r, column=WK_COST).value = label
        self.ws.row_dimensions[r].height = 24
        self.row += 1
        return r

    # -- header pair --------------------------------------------------------
    def headers(self, prev_week: date, mar31: date, *, exposure=True):
        """The two-row column header the client uses on every detail block."""
        ws, r = self.ws, self.row
        r2 = r + 1
        for rr in (r, r2):
            paint_row(ws, rr, 1, WK_LAST, font=F_HDR, fill=GOLD_DARK,
                      border=B_HDR, align=AL_CENTER_WRAP)

        def span(col, text, fmt=None, value2=None):
            ws.merge_cells(start_row=r, start_column=col, end_row=r2, end_column=col)
            style_cell(ws, r, col, text, font=F_HDR12)
            if value2 is not None:
                # A dated sub-header (e.g. the 31-Mar the column is measured from)
                # replaces the merge with a two-line pair.
                ws.unmerge_cells(start_row=r, start_column=col, end_row=r2, end_column=col)
                style_cell(ws, r2, col, value2, font=F_HDR, fmt=fmt)

        ws.merge_cells(start_row=r, start_column=WK_NAME, end_row=r2, end_column=WK_INCEPT)
        style_cell(ws, r, WK_NAME, "ASSET CLASS", font=F_HDR12)

        span(WK_PREV,  "Previous Week Value*", DATE_LONG, prev_week)
        span(WK_MAR31, "Market Value as on*",  DATE_LONG, mar31)
        span(WK_COST,  "COST*")
        if exposure:
            span(WK_EXP, "EXPOSURE (%)")
        span(WK_CMV,   "CURRENT MARKET VALUE*")
        span(WK_WKCHG, "WEEKLY CHANGE")

        ws.merge_cells(start_row=r, start_column=WK_PNL_YTD, end_row=r, end_column=WK_PNL_INC)
        style_cell(ws, r,  WK_PNL_YTD, "PROFIT & LOSS", font=F_HDR12)
        style_cell(ws, r2, WK_PNL_YTD, "YTD",        font=F_HDR)
        style_cell(ws, r2, WK_PNL_INC, "Inception",  font=F_HDR)

        span(WK_WKPCT, "WEEKLY CHANGE")

        ws.merge_cells(start_row=r, start_column=WK_RET_YTD, end_row=r, end_column=WK_RET_INC)
        style_cell(ws, r,  WK_RET_YTD, "RETURNS",   font=F_HDR12)
        style_cell(ws, r2, WK_RET_YTD, "YTD*",      font=F_HDR)
        style_cell(ws, r2, WK_RET_INC, "Inception", font=F_HDR)

        style_cell(ws, r2, WK_CAGR, "CAGR (Inception)", font=F_HDR7)
        span(WK_REMARK, "REMARKS")

        ws.row_dimensions[r].height  = 30
        ws.row_dimensions[r2].height = 18
        self.row += 2
        return r

    # -- body rows ----------------------------------------------------------
    def _write(self, vals, *, font, fill, border, money_fmt=MONEY, name_align=None):
        ws, r = self.ws, self.row
        paint_row(ws, r, 1, WK_LAST, font=font, fill=fill, border=border)
        for col, val in vals.items():
            c = style_cell(ws, r, col, val)
            if col == WK_NAME:
                c.alignment = name_align or AL_LEFT_WRAP
            elif col == WK_INCEPT:
                c.font = F_DATE_SM
                c.number_format = DATE_SHORT
                c.alignment = AL_CENTER
            elif col in WK_MONEY:
                c.number_format = money_fmt
                c.alignment = AL_RIGHT
            elif col in WK_PCT:
                c.number_format = PCT2
                c.alignment = AL_RIGHT
            elif col == WK_REMARK:
                c.font = F_REMARK
                c.alignment = AL_LEFT_WRAP
        self.row += 1
        return r

    def holding(self, h):
        """One live holding, in the client's plain 11pt body style."""
        cmv = _f(h, "current_value")
        vals = {
            WK_NAME:    _label_of(h),
            WK_INCEPT:  _inception_of(h),
            WK_PREV:    _f(h, "prev_week_value"),
            WK_MAR31:   _f(h, "market_value_as_on"),
            WK_COST:    _f(h, "cost"),
            WK_EXP:     (cmv / self.total_cmv) if (cmv and self.total_cmv) else None,
            WK_CMV:     cmv,
            WK_WKCHG:   _weekly_change(h),
            WK_PNL_YTD: _f(h, "pnl_ytd"),
            WK_PNL_INC: _f(h, "pnl_inception"),
            WK_WKPCT:   None,
            WK_RET_YTD: _pct(h, "returns_ytd_pct"),
            WK_RET_INC: _pct(h, "returns_inception_pct"),
            WK_CAGR:    _pct(h, "cagr_inception_pct"),
            WK_REMARK:  _remark_of(h),
        }
        prev = _f(h, "prev_week_value")
        if cmv is not None and prev:
            vals[WK_WKPCT] = (cmv - prev) / prev
        return self._write(vals, font=F_BODY, fill=NO_FILL, border=B_THIN)

    def _agg_vals(self, label, rows, remark=None):
        cmv  = _sum(rows, "current_value")
        prev = _sum(rows, "prev_week_value")
        cost = _sum(rows, "cost")
        pnl_ytd = _sum(rows, "pnl_ytd")
        pnl_inc = _sum(rows, "pnl_inception")
        wk, wk_base = _weekly_agg(rows)
        return {
            WK_NAME:    label,
            WK_PREV:    prev,
            WK_MAR31:   _sum(rows, "market_value_as_on"),
            WK_COST:    cost,
            WK_EXP:     (cmv / self.total_cmv) if (cmv and self.total_cmv) else None,
            WK_CMV:     cmv,
            WK_WKCHG:   wk,
            WK_PNL_YTD: pnl_ytd,
            WK_PNL_INC: pnl_inc,
            WK_WKPCT:   (wk / wk_base) if (wk is not None and wk_base) else None,
            WK_RET_YTD: (pnl_ytd / (cmv - pnl_ytd)) if (pnl_ytd is not None and cmv
                         and (cmv - pnl_ytd)) else None,
            WK_RET_INC: (pnl_inc / cost) if (pnl_inc is not None and cost) else None,
            WK_CAGR:    None,
            WK_REMARK:  remark,
        }

    def subtotal(self, label, rows, remark=None):
        """Light-gold bold sub-total, e.g. 'MF-LIQUID FUND TOTAL'."""
        return self._write(self._agg_vals(label, rows, remark),
                           font=F_SUBTOTAL, fill=GOLD_LIGHT, border=B_THIN,
                           money_fmt=MONEY_TOT, name_align=AL_LEFT)

    def grandtotal(self, label, rows, remark=None):
        """Dark-gold white-on-gold lettered total, e.g. 'A. TOTAL FIXED INVESMENT'."""
        return self._write(self._agg_vals(label, rows, remark),
                           font=F_GRANDTOT, fill=GOLD_DARK, border=B_BAND,
                           money_fmt=MONEY_TOT, name_align=AL_LEFT)

    def plain_total(self, label, vals, *, fill=GOLD_DARK, font=F_GRANDTOT):
        """A lettered row whose numbers are supplied rather than summed."""
        vals = dict(vals); vals[WK_NAME] = label
        return self._write(vals, font=font, fill=fill, border=B_BAND,
                           money_fmt=MONEY_TOT, name_align=AL_LEFT)

    def sub_label(self, text):
        """Un-filled bold group caption inside a block ('Equity Index Fund')."""
        r = self.row
        paint_row(self.ws, r, 1, WK_LAST, font=F_BODY, fill=NO_FILL, border=B_THIN)
        c = style_cell(self.ws, r, WK_NAME, text, font=F_BODY_BOLD, align=AL_LEFT)
        self.row += 1
        return r

    def footnote(self, text):
        r = self.row
        paint_row(self.ws, r, 1, WK_LAST, font=F_NOTE, fill=NO_FILL, align=AL_LEFT_WRAP)
        self.ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=WK_LAST)
        self.ws.cell(row=r, column=1).value = text
        self.row += 1
        return r


def _wk_summary_headers(w, prev_week: date, mar31: date):
    """Header pair for the Performance Summary block at the top of the page.

    Deliberately NOT the detail-block header: the client's summary carries a
    weekly-change PERCENT in K and YTD returns in L, where the detail block puts
    inception P&L and inception returns.
    """
    ws, r = w.ws, w.row
    r2 = r + 1
    for rr in (r, r2):
        paint_row(ws, rr, 1, WK_LAST, font=F_HDR, fill=GOLD_DARK,
                  border=B_HDR, align=AL_CENTER_WRAP)

    ws.merge_cells(start_row=r, start_column=WK_NAME, end_row=r2, end_column=WK_INCEPT)
    style_cell(ws, r, WK_NAME, "ASSET CLASS", font=F_HDR12)

    style_cell(ws, r,  WK_PREV,  "Previous Week Value", font=F_HDR)
    style_cell(ws, r2, WK_PREV,  prev_week, font=F_HDR, fmt=DATE_LONG)
    style_cell(ws, r,  WK_MAR31, "Market Value as on",  font=F_HDR)
    style_cell(ws, r2, WK_MAR31, mar31,     font=F_HDR, fmt=DATE_LONG)

    for col, text in ((WK_COST, "Cost / Value "), (WK_EXP, "EXPOSURE (%)"),
                      (WK_CMV, "CURRENT MARKET VALUE"), (WK_WKCHG, "WEEKLY CHANGE")):
        ws.merge_cells(start_row=r, start_column=col, end_row=r2, end_column=col)
        style_cell(ws, r, col, text, font=F_HDR12)

    # The summary carries no inception columns, so P&L and RETURNS are single
    # columns headed over a "YTD" qualifier, and weekly-change % sits between them.
    style_cell(ws, r,  WK_PNL_YTD,  "PROFIT & LOSS", font=F_HDR12)
    style_cell(ws, r2, WK_PNL_YTD,  "YTD", font=F_HDR)
    ws.merge_cells(start_row=r, start_column=WKS_WKPCT, end_row=r2,
                   end_column=WKS_WKPCT)
    style_cell(ws, r,  WKS_WKPCT,   "WEEKLY CHANGE", font=F_HDR12)
    style_cell(ws, r,  WKS_RET_YTD, "RETURNS", font=F_HDR12)
    style_cell(ws, r2, WKS_RET_YTD, "YTD", font=F_HDR)

    ws.row_dimensions[r].height  = 30
    ws.row_dimensions[r2].height = 18
    w.row += 2


def _wk_market_stats(w, benchmarks: list, mar31: date, prev_week: date):
    """The client's MARKET STATISTICS block: current / previous week / 31-Mar and
    the week and YTD percentage moves, one row per benchmark."""
    ws = w.ws
    r = w.row
    paint_row(ws, r, 1, WK_LAST, font=F_HDR12, fill=GOLD_DARK, border=B_BAND,
              align=AL_CENTER)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=WK_CMV)
    ws.cell(row=r, column=1).value = "MARKET STATISTICS"
    ws.cell(row=r, column=1).alignment = AL_LEFT
    ws.merge_cells(start_row=r, start_column=WK_PNL_YTD, end_row=r, end_column=WK_PNL_INC)
    ws.cell(row=r, column=WK_PNL_YTD).value = "RETURNS"
    w.row += 1

    r = w.row
    paint_row(ws, r, 1, WK_LAST, font=F_HDR, fill=GOLD_DARK, border=B_HDR,
              align=AL_CENTER_WRAP)
    style_cell(ws, r, WK_EXP,     "Current",       font=F_HDR)
    style_cell(ws, r, WK_CMV,     "Previous Week", font=F_HDR)
    style_cell(ws, r, WK_WKCHG,   mar31,           font=F_HDR, fmt=DATE_LONG)
    style_cell(ws, r, WK_PNL_YTD, "WEEKLY %",      font=F_HDR)
    style_cell(ws, r, WK_PNL_INC, "YTD %",         font=F_HDR)
    ws.row_dimensions[r].height = 20
    w.row += 1

    by_code = {b["code"]: b for b in (benchmarks or [])}
    for code, label in BENCHMARK_ROWS:
        b = by_code.get(code) or {}
        r = w.row
        paint_row(ws, r, 1, WK_LAST, font=F_BODY, fill=NO_FILL, border=B_THIN)
        style_cell(ws, r, WK_NAME, label, font=F_BODY_BOLD, align=AL_LEFT)
        ws.merge_cells(start_row=r, start_column=WK_NAME, end_row=r, end_column=WK_INCEPT)
        style_cell(ws, r, WK_EXP,     b.get("current"),   fmt=PRICE, align=AL_RIGHT)
        style_cell(ws, r, WK_CMV,     b.get("prev_week"), fmt=PRICE, align=AL_RIGHT)
        style_cell(ws, r, WK_WKCHG,   b.get("mar31"),     fmt=PRICE, align=AL_RIGHT)
        style_cell(ws, r, WK_PNL_YTD, b.get("week_pct"),  fmt=PCT2,  align=AL_RIGHT)
        style_cell(ws, r, WK_PNL_INC, b.get("ytd_pct"),   fmt=PCT2,  align=AL_RIGHT)
        w.row += 1

    w.footnote("ADD NI= New Issue")


# Sub-groups inside "MF EQUITY TOTAL (i)", in the client's own order and wording.
_MF_EQUITY_GROUPS = [
    ("Equity Index Fund",         "mf_index"),
    ("Market Equity Fund",        "mf_market"),
    ("Sector/Thematic Equity Fund", "mf_sector"),
    ("Hybrid (Debt & Equity) Fund", "mf_hybrid"),
]


def _realised_row(w, realised_total, label="Realised Profit & Loss (in YTD)"):
    """The client shows realised P&L as its own line under each asset class. The
    figure is already inside the class total above it — this line only breaks it out
    (the client's own note at the bottom of the all-entities sheet says so)."""
    return w._write({WK_NAME: label, WK_PNL_YTD: realised_total},
                    font=F_BODY, fill=NO_FILL, border=B_THIN, name_align=AL_LEFT)


def build_weekly_report(ws, label: str, bundle: dict, as_of: date,
                        benchmarks: list = None, realised: list = None,
                        staleness: str = None):
    """
    Render one entity's / group's weekly page in the client's exact layout.

    Blocks, in the client's order:
      title → Performance Summary → Market Statistics
      → '- FIXED INCOME INVESTMENT'  (liquid / debt / arbitrage / PPF → A. total)
      → '- EQUITIES'                 (MF / PMS / direct / foreign / AIF → B, C totals)
      → '- ALTERNATES'               (unlisted / forex / gold → D, then E-G, H, H1)
    """
    ws.sheet_view.showGridLines = False
    set_widths(ws, WK_WIDTHS)

    sec    = classify_bundle(bundle)
    totals = section_totals(sec)
    total_cmv = totals["grand"]["cmv"] or 0.0
    mar31     = _mar31(as_of)
    prev_week = date.fromordinal(as_of.toordinal() - 7)

    # Realised P&L bucketed the way the client's asset-class lines are.
    rl = defaultdict(float)
    for r in (realised or []):
        g = (r.get("group") or "").lower()
        pnl = _f(r, "pnl") or 0.0
        if g in ("fixed income", "mutual funds"):
            rl["fixed"] += pnl
        elif g in ("commodities", "alternates"):
            rl["alt"] += pnl
        else:
            rl["equity"] += pnl
    rl_total = sum(rl.values()) if rl else None

    w = _WeeklyWriter(ws, as_of, total_cmv)

    # ── page title + performance summary ──────────────────────────────────────
    w.title(f"ALL Assets Monitor - {label.upper()}")

    r = w.row
    paint_row(ws, r, 1, WK_LAST, font=F_BANNER20, fill=GOLD_DARK, border=B_BAND,
              align=AL_CENTER)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=WK_CMV)
    c = ws.cell(row=r, column=1)
    c.value = f"Performance Summary For {label.upper()} As on "
    c.alignment = AL_RIGHT
    ws.merge_cells(start_row=r, start_column=WK_WKCHG, end_row=r, end_column=WK_LAST)
    d = ws.cell(row=r, column=WK_WKCHG)
    d.value = as_of
    d.number_format = r'[$-409]mmmm\ d\,\ yyyy'
    d.alignment = AL_LEFT
    ws.row_dimensions[r].height = 26
    w.row += 1

    _wk_summary_headers(w, prev_week, mar31)

    def summary_line(name, t, fill=NO_FILL, font=F_BODY_BOLD):
        cmv = t["cmv"]
        w._write({
            WK_NAME: name, WK_PREV: t["prev"], WK_MAR31: t["mar31"],
            WK_COST: t["cost"],
            WK_EXP: (cmv / total_cmv) if (cmv and total_cmv) else None,
            WK_CMV: cmv, WK_WKCHG: t["wk"], WK_PNL_YTD: t["pnl_ytd"],
            WKS_WKPCT: t["wk_pct"], WKS_RET_YTD: t["ret_ytd"],
        }, font=font, fill=fill, border=B_THIN, money_fmt=MONEY, name_align=AL_LEFT)

    summary_line("Fixed Income",    totals["fixed"])
    summary_line("Equity ",         totals["equity"])
    summary_line("Alternates",      totals["alt"])
    summary_line("Funds In Transit", totals["transit"])
    summary_line("Broker Balance",  totals["broker"])
    summary_line("Funds In Bank",   totals["bank"])
    summary_line("GRAND TOTAL",     totals["grand"], fill=GOLD_DARK, font=F_GRANDTOT)

    w.blank()
    _wk_market_stats(w, benchmarks, mar31, prev_week)
    w.blank()

    # ── FIXED INCOME ──────────────────────────────────────────────────────────
    w.title("- FIXED INCOME INVESTMENT")
    w.entity_banner(label.upper())
    w.headers(prev_week, mar31)

    w.subtotal("MF-LIQUID FUND TOTAL", sec["liquid"], "Weighted Avg Return % ")
    for h in sec["liquid"]:
        w.holding(h)
    _realised_row(w, None)
    w.footnote("*YTD returns on liquid funds have been shown as XIRR due to the larger "
               "number of additions and redemptions")

    w.subtotal("DEBT TOTAL", sec["debt"], "Annualised YTD return %  ")
    for h in sec["debt"]:
        w.holding(h)
    _realised_row(w, None)
    w.blank()

    w.subtotal("ARBITRAGE FUND", sec["arbitrage"])
    for h in sec["arbitrage"]:
        w.holding(h)
    _realised_row(w, None)
    w.blank()

    w.subtotal("PUBLIC PROVIDENT FUND", sec["ppf"], "Annualised YTD return %")
    for h in sec["ppf"]:
        w.holding(h)
    _realised_row(w, None)
    w.blank()

    w.grandtotal("A. TOTAL FIXED INVESMENT", totals["fixed"]["rows"],
                 "Weighted Avg Return %")
    _realised_row(w, rl.get("fixed") or None)
    w.footnote("*Any additional input values like SIP payments or Switch Ins since 1 april "
               "has been also added to the total value as on 1 april and redemptions have "
               "been reduced ")
    w.blank()

    # ── EQUITIES ──────────────────────────────────────────────────────────────
    w.title("- EQUITIES")
    w.entity_banner(label.upper())
    w.headers(prev_week, mar31)

    mf_equity = _rows_of(sec, ["mf_index", "mf_market", "mf_sector", "mf_hybrid"])
    w.subtotal("MF EQUITY TOTAL (i)", mf_equity, "All absolute %")
    for caption, key in _MF_EQUITY_GROUPS:
        if sec[key]:
            w.sub_label(caption)
            for h in sec[key]:
                w.holding(h)
    _realised_row(w, None)
    w.blank()

    w.subtotal("PMS EQUITY TOTAL (ii)", sec["pms"])
    for h in sec["pms"]:
        w.holding(h)
    _realised_row(w, None)
    w.blank()

    w.subtotal("DIRECT EQUITIES DESK (iii)", sec["direct"])
    for h in sec["direct"]:
        w.holding(h)
    _realised_row(w, None, "Realised Profit & Loss +  Dividends (in YTD)")
    w.blank()

    w.subtotal("FOREIGN EQUITY & FUNDS (iv)", sec["foreign"])
    for h in sec["foreign"]:
        w.holding(h)
    _realised_row(w, None)
    w.blank()

    w.subtotal("ALTERNATE INVESTMENT FUND TOTAL", sec["aif"])
    for h in sec["aif"]:
        w.holding(h)
    _realised_row(w, None)
    w.blank()

    w.grandtotal("B. TOTAL EQUITY EXPOSURE (i+ii+iii+iv)", totals["equity"]["rows"])
    _realised_row(w, rl.get("equity") or None)
    w.blank()
    w.grandtotal("C. SUB GRAND TOTAL (A+B)", totals["sub"]["rows"], "Weighted Av Return %")
    w.footnote("*Any additional input values like SIP payments or Switch Ins since 1 april "
               "has been also added to the total value as on 1 april and redemptions have "
               "been reduced ")
    w.blank()

    # ── ALTERNATES ────────────────────────────────────────────────────────────
    w.title("- ALTERNATES")
    w.entity_banner(label.upper())
    w.headers(prev_week, mar31)

    w.subtotal("ALTERNATE EQUITIES ", sec["unlisted"], "All absolute %")
    for h in sec["unlisted"]:
        w.holding(h)
    _realised_row(w, None)
    w.blank()

    w.subtotal("FOREX TOTAL", sec["forex"])
    for h in sec["forex"]:
        w.holding(h)
    _realised_row(w, None)
    w.blank()

    w.subtotal("GOLD ETF TOTAL", sec["gold"])
    for h in sec["gold"]:
        w.holding(h)
    _realised_row(w, rl.get("alt") or None)
    w.blank()

    w.grandtotal("D. Alternates Total", totals["alt"]["rows"])
    w.blank()

    w.subtotal("E. Funds In transit", sec["transit"])
    for h in sec["transit"]:
        w.holding(h)
    w.subtotal("F. Broker Balance", sec["broker"])
    for h in sec["broker"]:
        w.holding(h)
    w.subtotal("G. Funds in Bank", sec["bank"])
    for h in sec["bank"]:
        w.holding(h)
    w.blank()

    w.grandtotal("H. Grand Total (C+D+E+F+G)", totals["grand"]["rows"])
    w.blank()
    ex_forex = [h for h in (totals["sub"]["rows"] + totals["alt"]["rows"])
                if h not in sec["forex"]]
    w.grandtotal("H1. Grand Total (C+D)-FOREX", ex_forex)

    w.footnote("*Weekly Change column includes change in Value from previous week and "
               "does not represent returns")
    w.footnote("*YTD of Fixed Income Funds Invested less than a year ago taken from "
               "Money Control")
    # A feed that has not refreshed makes every number above it quietly stale. The
    # client's format has no cell for that, so it rides along as one more footnote
    # rather than being dropped — a silently old report is the worse failure.
    if staleness:
        w.footnote(staleness)

    # Weighted benchmarks the client tracks against the MF-Equity and PMS sleeves.
    by_code = {b["code"]: b for b in (benchmarks or [])}
    for series, sleeve in (("NIFTY", "MF-Equity"), ("SENSEX", "MF-Equity"),
                           ("NIFTY", "PMS"), ("SENSEX", "PMS")):
        b = by_code.get(series) or {}
        name = f"Weighted {'NSE - Nifty' if series == 'NIFTY' else 'BSE - Sensex'} ({sleeve})"
        r = w.row
        paint_row(ws, r, 1, WK_LAST, font=F_BODY, fill=NO_FILL, border=B_THIN)
        style_cell(ws, r, WK_NAME, name, font=F_BODY, align=AL_LEFT)
        ws.merge_cells(start_row=r, start_column=WK_NAME, end_row=r, end_column=WK_INCEPT)
        style_cell(ws, r, WK_EXP,   b.get("current"),   fmt=PRICE, align=AL_RIGHT)
        style_cell(ws, r, WK_CMV,   b.get("prev_week"), fmt=PRICE, align=AL_RIGHT)
        style_cell(ws, r, WK_WKCHG, b.get("mar31"),     fmt=PRICE, align=AL_RIGHT)
        style_cell(ws, r, WK_PNL_YTD, b.get("week_pct"), fmt=PCT2, align=AL_RIGHT)
        style_cell(ws, r, WK_WKPCT,   b.get("ytd_pct"),  fmt=PCT2, align=AL_RIGHT)
        w.row += 1

    ws.freeze_panes = "A5"
    ws.print_title_rows = "1:4"
    return ws


# ══════════════════════════════════════════════════════════════════════════════
#  Equity Daily Print  —  one 13-column block per entity, domestic then foreign
# ══════════════════════════════════════════════════════════════════════════════

EDP_LAST = 13
EDP_WIDTHS = {
    "A": 31.57, "B": 15.00, "C": 13.86, "D": 8.71, "E": 16.00, "F": 15.00,
    "G": 14.00, "H": 14.00, "I": 15.43, "J": 12.86, "K": 12.71, "L": 11.57,
    "M": 14.14,
}

# The client's four header wordings — a group-total block says "Average Inception
# Returns %" where an entity block says "Inception Returns %", and the foreign
# blocks carry "($)". Reproduced verbatim, typos included, so the sheet reads
# exactly as the client's own.
def _edp_headers(total: bool, foreign: bool) -> list:
    """The client's header wording for this block variant, reproduced verbatim.

    The four variants differ in small ways that are inconsistent in the client's own
    file — "Script Name" on a group total but "Scrip Name" on an entity block,
    "Inception" spelled "Incetion" on entity blocks, a capital E in "Stock Exposure
    (Cost)" only on the foreign blocks. They are copied as-is rather than tidied:
    matching the client's sheet is the requirement.
    """
    d = " ($)" if foreign else ""
    return [
        # Only the DOMESTIC group total is headed "Script Name"; every other block —
        # including the foreign group total — says "Scrip Name" in the client's file.
        "Script Name" if (total and not foreign) else "Scrip Name",
        "Ticker Name",
        "Last Purchase Date",
        "Quantity",
        f"{'Average' if total else 'AVERAGE'} Purchase Price{d}",
        "Total Purchase Cost",
        f"Current Market Price{d}",
        f"Total Current Market Value{d} (HIGHEST TO LOWEST SORTED)",
        f"Total Profit/Loss (From {'Inception' if total else 'Incetion'}){d}",
        f"{'Average ' if total else ''}Inception Returns % (Absolute)",
        "Stock Exposure (CMP) %",
        f"Stock {'Exposure' if foreign else 'exposure'} (Cost) %",
        "Change in Stock Exposure (CMP - Cost)",
    ]


def _edp_block(ws, row: int, title: str, as_of: date, holdings: list,
               *, total: bool, foreign: bool) -> int:
    """One title / header / rows / Grand Total block. Returns the next free row."""
    # Title bar — label on the left, the as-on date on the right, both pale gold.
    paint_row(ws, row, 1, EDP_LAST, font=F_TITLE16, fill=GOLD_PALE, border=B_BAND,
              align=AL_CENTER_WRAP)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=EDP_LAST)
    t = ws.cell(row=row, column=1)
    t.value = title
    t.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    d = ws.cell(row=row, column=8)
    d.value = as_of
    d.number_format = r'[$-409]mmmm\ d\,\ yyyy'
    d.alignment = AL_CENTER
    ws.row_dimensions[row].height = 24
    row += 1

    # Column headers — amber, bold, wrapped.
    paint_row(ws, row, 1, EDP_LAST, font=F_HDR_DARK, fill=AMBER, border=B_HDR,
              align=AL_CENTER_WRAP)
    for c, text in enumerate(_edp_headers(total, foreign), 1):
        ws.cell(row=row, column=c).value = text
    ws.row_dimensions[row].height = 45
    row += 1

    tot_cmv  = _sum(holdings, "current_market_value") or 0.0
    tot_cost = _sum(holdings, "cost") or 0.0

    for h in holdings:
        cmv  = _f(h, "current_market_value")
        cost = _f(h, "cost")
        exp_cmp  = (cmv / tot_cmv) if (cmv and tot_cmv) else None
        exp_cost = (cost / tot_cost) if (cost and tot_cost) else None
        vals = [
            h.get("security_name") or h.get("symbol"),
            h.get("symbol"),
            h.get("last_purchase_date") or _inception_of(h),
            _f(h, "quantity"),
            _f(h, "avg_cost"),
            cost,
            _f(h, "current_price"),
            cmv,
            _f(h, "pnl_inception"),
            _pct(h, "returns_inception_pct"),
            exp_cmp,
            exp_cost,
            (exp_cmp - exp_cost) if (exp_cmp is not None and exp_cost is not None) else None,
        ]
        paint_row(ws, row, 1, EDP_LAST, font=F_BODY, fill=NO_FILL, border=B_THIN)
        for c, v in enumerate(vals, 1):
            cell = style_cell(ws, row, c, v)
            if c == 3:
                cell.number_format = DATE_DMY; cell.alignment = AL_CENTER
            elif c == 4:
                cell.number_format = QTY;      cell.alignment = AL_RIGHT
            elif c in (5, 7):
                cell.number_format = PRICE;    cell.alignment = AL_RIGHT
            elif c in (6, 8, 9):
                cell.number_format = MONEY;    cell.alignment = AL_RIGHT
            elif c >= 10:
                cell.number_format = PCT2;     cell.alignment = AL_RIGHT
            else:
                cell.alignment = AL_LEFT_WRAP
        row += 1

    # Grand Total — pale blue, bold, double rule above.
    paint_row(ws, row, 1, EDP_LAST, font=F_SUBTOTAL, fill=PALE_BLUE, border=B_TOTAL,
              align=AL_RIGHT)
    gt = [
        "Grand Total", None,
        max([h.get("last_purchase_date") or _inception_of(h) for h in holdings
             if (h.get("last_purchase_date") or _inception_of(h))], default=None),
        _sum(holdings, "quantity"), None, tot_cost or None, None,
        tot_cmv or None, _sum(holdings, "pnl_inception"),
        ((_sum(holdings, "pnl_inception") / tot_cost) if tot_cost else None),
        (1.0 if tot_cmv else None), (1.0 if tot_cost else None), None,
    ]
    for c, v in enumerate(gt, 1):
        cell = style_cell(ws, row, c, v)
        if c == 1:
            cell.alignment = AL_LEFT
        elif c == 3:
            cell.number_format = DATE_DMY; cell.alignment = AL_CENTER
        elif c == 4:
            cell.number_format = QTY
        elif c in (6, 8, 9):
            cell.number_format = MONEY_TOT
        elif c >= 10:
            cell.number_format = PCT2
    ws.row_dimensions[row].height = 18
    return row + 2          # one blank row between blocks, as the client has


def build_equity_daily_print(ws, as_of: date, group_label: str,
                             domestic_by_entity: dict, foreign_by_entity: dict):
    """
    The client's Equity Daily Print tab: a whole-group DOMESTIC block, one block per
    entity holding domestic equity, then the same for FOREIGN, closing with the
    whole-group foreign block.

    domestic_by_entity / foreign_by_entity: {entity_code: [holding, ...]}, each list
    already sorted highest current market value first (the header says so).
    """
    ws.sheet_view.showGridLines = False
    set_widths(ws, EDP_WIDTHS)
    row = 1

    all_dom = [h for rows in domestic_by_entity.values() for h in rows]
    all_for = [h for rows in foreign_by_entity.values() for h in rows]

    if all_dom:
        row = _edp_block(ws, row, f"Total {group_label} Direct Equity (DOMESTIC) as on",
                         as_of, all_dom, total=True, foreign=False)
    for code, rows in domestic_by_entity.items():
        if rows:
            row = _edp_block(ws, row, f"{code} Direct Equity (DOMESTIC) as on",
                             as_of, rows, total=False, foreign=False)
    for code, rows in foreign_by_entity.items():
        if rows:
            row = _edp_block(ws, row, f"{code} Direct Equity (FOREIGN) ($) as on",
                             as_of, rows, total=False, foreign=True)
    if all_for:
        row = _edp_block(ws, row, f"Total {group_label} Direct Equity (FOREIGN) ($) as on",
                         as_of, all_for, total=True, foreign=True)
    return ws


# ══════════════════════════════════════════════════════════════════════════════
#  All Assets Daily MIS  —  benchmarks, then Fixed Income / Equity / Alternates
# ══════════════════════════════════════════════════════════════════════════════

AA_LAST = 8
AA_WIDTHS = {"A": 75.57, "B": 17.57, "C": 12.86, "D": 16.57, "E": 15.71,
             "F": 15.14, "G": 12.29, "H": 12.14}

# Column B names where the number comes from — the client's own operating note on
# which portal or statement has to be opened to refresh that line.
SRC_MF    = "CAMS / KARVY"
SRC_LOGIN = "LOGIN"
SRC_BANK  = "BANK"
SRC_DEMAT = "DEMAT LOGIN"


class _AllAssetsWriter:
    def __init__(self, ws, as_of):
        self.ws = ws
        self.row = 1
        self.as_of = as_of

    def blank(self, n=1):
        self.row += n

    def section_header(self, title, mar31, prev_week, weekly_label, ytd_label):
        """The amber two-row band that opens Fixed Income / Equity / Alternates."""
        ws, r = self.ws, self.row
        r2 = r + 1
        for rr in (r, r2):
            paint_row(ws, rr, 1, AA_LAST, font=F_HDR_DARK, fill=AMBER,
                      border=B_HDR, align=AL_CENTER_WRAP)
        ws.merge_cells(start_row=r, start_column=1, end_row=r2, end_column=1)
        style_cell(ws, r, 1, title, font=F_HDR_DARK, align=AL_CENTER_WRAP)
        ws.merge_cells(start_row=r, start_column=3, end_row=r2, end_column=3)
        style_cell(ws, r, 3, "Original Cost")
        style_cell(ws, r, 4, "Market Value on")
        style_cell(ws, r2, 4, mar31, fmt=r'd" "mmm" "yyyy')
        ws.merge_cells(start_row=r, start_column=5, end_row=r2, end_column=5)
        style_cell(ws, r, 5, "Current Market Value")
        style_cell(ws, r, 6, "Previous Week")
        style_cell(ws, r2, 6, prev_week, fmt=r'd" "mmm" "yyyy')
        ws.merge_cells(start_row=r, start_column=7, end_row=r2, end_column=7)
        style_cell(ws, r, 7, weekly_label)
        ws.merge_cells(start_row=r, start_column=8, end_row=r2, end_column=8)
        style_cell(ws, r, 8, ytd_label)
        ws.row_dimensions[r].height = 30
        self.row += 2

    def caption(self, text, source=None):
        """Un-totalled group caption, e.g. 'MF-Liquid Fund'."""
        r = self.row
        paint_row(self.ws, r, 1, AA_LAST, font=F_BODY12, fill=NO_FILL, border=B_THIN)
        style_cell(self.ws, r, 1, text, font=F_BODY_BOLD, align=AL_LEFT)
        if source:
            style_cell(self.ws, r, 2, source, font=F_BODY12, align=AL_CENTER)
        self.row += 1

    def _values(self, r, cost, mar31, cmv, prev, ytd_pct, fmt=MONEY, wk=None):
        if wk is None and cmv is not None and prev is not None:
            wk = cmv - prev
        for col, val, f in ((3, cost, fmt), (4, mar31, fmt), (5, cmv, fmt),
                            (6, prev, fmt), (7, wk, fmt), (8, ytd_pct, PCT2)):
            style_cell(self.ws, r, col, val, fmt=f, align=AL_RIGHT)

    def holding(self, h, source):
        r = self.row
        paint_row(self.ws, r, 1, AA_LAST, font=F_BODY12, fill=NO_FILL, border=B_THIN)
        style_cell(self.ws, r, 1, _label_of(h), font=F_BODY12, align=AL_LEFT)
        style_cell(self.ws, r, 2, source, font=F_BODY12, align=AL_CENTER)
        self._values(r, _f(h, "cost"), _f(h, "market_value_as_on"),
                     _f(h, "current_value"), _f(h, "prev_week_value"),
                     _pct(h, "returns_ytd_pct"))
        self.row += 1

    def total(self, text, rows, *, fill=CREAM, font=F_SECTION):
        r = self.row
        paint_row(self.ws, r, 1, AA_LAST, font=font, fill=fill, border=B_THIN)
        style_cell(self.ws, r, 1, text, font=font, align=AL_LEFT)
        cmv     = _sum(rows, "current_value")
        pnl_ytd = _sum(rows, "pnl_ytd")
        wk, _base = _weekly_agg(rows)
        self._values(r, _sum(rows, "cost"), _sum(rows, "market_value_as_on"), cmv,
                     _sum(rows, "prev_week_value"),
                     (pnl_ytd / (cmv - pnl_ytd)) if (pnl_ytd is not None and cmv
                                                     and (cmv - pnl_ytd)) else None,
                     fmt=MONEY_TOT, wk=wk)
        self.row += 1

    def group(self, caption, rows, total_label, source, sub_captions=None):
        """Caption → holdings (optionally sub-captioned) → the client's lettered total."""
        if caption:
            self.caption(caption, source)
        if sub_captions:
            for sub, sub_rows in sub_captions:
                if sub_rows:
                    self.caption(sub, source)
                    for h in sub_rows:
                        self.holding(h, source)
        else:
            for h in rows:
                self.holding(h, source)
        self.total(total_label, rows)


def build_all_assets_daily(ws, as_of: date, sec: dict, benchmarks: list = None):
    """The client's All Assets Daily MIS tab, driven by one combined all-entity bundle."""
    ws.sheet_view.showGridLines = False
    set_widths(ws, AA_WIDTHS)
    mar31     = _mar31(as_of)
    prev_week = date.fromordinal(as_of.toordinal() - 7)
    fy        = _fy_label(as_of)
    w = _AllAssetsWriter(ws, as_of)

    # ── benchmark strip ───────────────────────────────────────────────────────
    r = w.row
    paint_row(ws, r, 1, AA_LAST, font=F_TITLE, fill=GOLD_PALE, border=B_THIN,
              align=AL_LEFT)
    style_cell(ws, r, 1, "ALL ASSETS DAILY REPORT")
    style_cell(ws, r, 2, "Data")
    ws.row_dimensions[r].height = 30
    w.row += 1

    r = w.row
    paint_row(ws, r, 1, AA_LAST, font=F_SECTION, fill=CREAM, border=B_THIN,
              align=AL_CENTER)
    style_cell(ws, r, 1, "Benchmarks", align=AL_LEFT)
    for col, text in ((3, "Current"), (4, "Previous Week"), (5, "Yesterday"),
                      (6, f"FY {fy}"), (7, "Daily"), (8, "Weekly")):
        style_cell(ws, r, col, text)
    w.row += 1

    by_code = {b["code"]: b for b in (benchmarks or [])}
    for code, label in (("NIFTY", "Nifty"), ("SENSEX", "Sensex")):
        b = by_code.get(code) or {}
        r = w.row
        paint_row(ws, r, 1, AA_LAST, font=F_BODY12, fill=NO_FILL, border=B_THIN)
        style_cell(ws, r, 1, label, font=F_BODY_BOLD, align=AL_LEFT)
        style_cell(ws, r, 2, "LIVE", align=AL_CENTER)
        style_cell(ws, r, 3, b.get("current"),    fmt=PRICE, align=AL_RIGHT)
        style_cell(ws, r, 4, b.get("prev_week"),  fmt=PRICE, align=AL_RIGHT)
        style_cell(ws, r, 5, b.get("prev_close"), fmt=PRICE, align=AL_RIGHT)
        style_cell(ws, r, 6, b.get("mar31"),      fmt=PRICE, align=AL_RIGHT)
        style_cell(ws, r, 7, b.get("day_pct"),    fmt=PCT2,  align=AL_RIGHT)
        style_cell(ws, r, 8, b.get("week_pct"),   fmt=PCT2,  align=AL_RIGHT)
        w.row += 1

    w.blank(2)
    r = w.row
    paint_row(ws, r, 1, AA_LAST, font=F_TITLE16, fill=NO_FILL, border=B_THIN,
              align=AL_CENTER)
    style_cell(ws, r, 1, f"Asset Class Returns_F.Y. {fy}")
    w.row += 1

    # ── Fixed Income ──────────────────────────────────────────────────────────
    w.section_header("Fixed Income Investments", mar31, prev_week,
                     "Weekly Change", "YTD Annualised")
    w.group("MF-Liquid Fund",   sec["liquid"],    "Total MF Liquid Fund (A)",   SRC_MF)
    w.blank()
    w.group("MF - Debt Fund",   sec["debt"],      "Total MF- Debt Fund (B)",    SRC_MF)
    w.blank()
    w.group("MF- Arbitrage Fund", sec["arbitrage"], "Total MF- Arbitrage Fund (C)", SRC_MF)
    w.blank()
    w.group("MF - PPF",         sec["ppf"],       "Total PPF (D)",              SRC_BANK)
    w.blank()

    # ── Equity ────────────────────────────────────────────────────────────────
    w.section_header("Equity Investment", mar31, prev_week, "Wkly Chg ", "YTD Absolute")
    mf_equity = _rows_of(sec, ["mf_index", "mf_market", "mf_sector", "mf_hybrid"])
    w.group("MF-Equity Funds", mf_equity, "Total MF Equity (E)", SRC_MF,
            sub_captions=[("Equity Index Fund",          sec["mf_index"]),
                          ("Market Equity Fund",         sec["mf_market"]),
                          ("Sector/Thematic Equity Fund", sec["mf_sector"]),
                          ("Hybrid (Debt & Equity) Fund", sec["mf_hybrid"])])
    w.blank()
    w.group("PMS ",          sec["pms"],     "Total Equity PMS (F)",    SRC_LOGIN)
    w.blank()
    w.group("Direct Stocks", sec["direct"],  "Total Equity Direct (G)", SRC_LOGIN)
    w.blank()
    w.group("Alternate Investment Fund", sec["aif"], "Total AIF Equity (G)", SRC_LOGIN)
    w.blank()
    w.total("Total Equity (E+F+G) ",
            mf_equity + sec["pms"] + sec["direct"] + sec["aif"])
    w.blank()

    # ── Alternates ────────────────────────────────────────────────────────────
    w.section_header("Alternate Investments", mar31, prev_week,
                     "Weekly Change ", "YTD Absolute")
    w.group("Overseas Direct Equity", sec["foreign"], "Forex Total", SRC_LOGIN)
    w.blank()
    w.group("Alternate Indian Equity", sec["unlisted"],
            "Total Alternate Indian Equity", SRC_LOGIN)
    w.blank()
    w.group("GOLD/ SILVER ETF", sec["gold"], "Total Gold / Silver ETF", SRC_DEMAT)
    w.blank()
    w.caption("Funds in Transit")
    for h in sec["transit"]:
        w.holding(h, SRC_BANK)
    w.caption("Broker Balance", SRC_DEMAT)
    for h in sec["broker"]:
        w.holding(h, SRC_DEMAT)
    w.caption("Funds in Bank", SRC_BANK)
    for h in sec["bank"] + sec["forex"]:
        w.holding(h, SRC_BANK)
    w.blank()

    # ── grand totals ──────────────────────────────────────────────────────────
    # "Non-earning assets" are the idle balances: cash in transit, broker float and
    # bank/forex balances. The client reports the book both ways, so a reader can
    # see returns on invested capital without the drag of uninvested cash.
    earning     = (_rows_of(sec, FIXED_BUCKETS) + _rows_of(sec, EQUITY_BUCKETS)
                   + _rows_of(sec, ALT_BUCKETS))
    non_earning = _rows_of(sec, BELOW_BUCKETS)
    alt_rows    = _rows_of(sec, ALT_BUCKETS)

    w.total("Total Alternates (H) (including non earning Assets)",
            alt_rows + non_earning, fill=CREAM)
    w.total("Total Alternates (H1) (excluding non earning Assets)", alt_rows, fill=CREAM)
    w.total("Grand Total (A to H) (including non earning Assets)",
            earning + non_earning, fill=GOLD_LIGHT)
    w.total("Grand Total (A to H1) (excluding non earning Assets)", earning,
            fill=GOLD_LIGHT)
    note(ws, w.row, 1,
         "*Weekly Change column includes change in Value from previous week and does "
         "not represent returns", span=AA_LAST)
    return ws


# ══════════════════════════════════════════════════════════════════════════════
#  All Entities Weekly Report  —  5 columns per group + an ALL RAJANI total
# ══════════════════════════════════════════════════════════════════════════════

# Per-group column headers, in the client's order and wording.
AE_GROUP_COLS = ["CMV", "Weekly Change", "Profit & Loss", "Weekly Change %", "Returns (%)"]
# Trailing manual-annotation columns the client keeps (three of them hidden).
AE_TAIL_COLS  = [("Total - Avg Invt", True), ("GKP", True), ("AP", True),
                 ("AKP", True), ("Remarks", False)]

# The lettered lines down column A, and which bucket(s) each one sums.
AE_LINES = [
    ("MF-LIQUID FUND",                  ["liquid"]),
    ("MF-DEBT FUND",                    ["debt"]),
    ("ARBITRAGE FUND",                  ["arbitrage"]),
    ("PPF",                             ["ppf"]),
    ("Realised Profit & Loss",          "realised:fixed"),
    ("A. TOTAL FIXED INCOME INVESMENT", FIXED_BUCKETS),
    (None, None),
    ("MF EQUITY",                       ["mf_index", "mf_market", "mf_sector", "mf_hybrid"]),
    ("PMS EQUITY ",                     ["pms"]),
    ("DIRECT DESK",                     ["direct"]),
    ("OVERSEAS INVESTMENTS",            ["foreign"]),
    ("ALTERNATE INVESTMENT FUND ",      ["aif"]),
    ("Realised Profit & Loss",          "realised:equity"),
    ("B. TOTAL EQUITY ",                EQUITY_BUCKETS),
    (None, None),
    ("C. SUB GRAND TOTAL (A+B)",        FIXED_BUCKETS + EQUITY_BUCKETS),
    (None, None),
    ("STARTUPS & UNLISTED EQUITY",      ["unlisted"]),
    ("GOLD/ SILVER ETF",                ["gold"]),
    ("FOREX TOTAL",                     ["forex"]),
    ("Realised Profit & Loss",          "realised:alt"),
    (None, None),
    ("D. ALTERNATES",                   ALT_BUCKETS),
    (None, None),
    ("E. FUNDS IN TRANSIT",             ["transit"]),
    ("F. BROKER BALANCE",               ["broker"]),
    ("G. FUNDS IN BANK",                ["bank"]),
    (None, None),
    ("H1. GRAND TOTAL*",                FIXED_BUCKETS + EQUITY_BUCKETS
                                        + ALT_BUCKETS + BELOW_BUCKETS),
]

# Lines the client renders as a filled total bar rather than a plain row.
_AE_TOTAL_LINES = {"A. TOTAL FIXED INCOME INVESMENT", "B. TOTAL EQUITY ",
                   "C. SUB GRAND TOTAL (A+B)", "D. ALTERNATES", "H1. GRAND TOTAL*"}


def build_all_entities_weekly(ws, as_of: date, group_blocks: list,
                              benchmarks: list = None):
    """
    The client's ASSET MONITOR: one 5-column block per PAN group, closing with an
    ALL RAJANI GROUP block that totals every group.

    group_blocks: [{"label": "Dhruv GROUP", "sec": <classify_bundle result>,
                    "realised": {"fixed": x, "equity": y, "alt": z}}, ...]
    in the order the client wants them across the sheet.
    """
    ws.sheet_view.showGridLines = False
    mar31     = _mar31(as_of)
    prev_week = date.fromordinal(as_of.toordinal() - 7)

    # An ALL RAJANI GROUP block whose buckets are every group's buckets concatenated.
    combined_sec = defaultdict(list)
    combined_rl  = defaultdict(float)
    for g in group_blocks:
        for k, v in g["sec"].items():
            combined_sec[k].extend(v)
        for k, v in (g.get("realised") or {}).items():
            combined_rl[k] += v or 0.0
    blocks = list(group_blocks) + [{"label": "ALL RAJANI GROUP",
                                    "sec": combined_sec, "realised": combined_rl}]

    n_blocks = len(blocks)
    last_val_col = 1 + n_blocks * 5
    last_col     = last_val_col + len(AE_TAIL_COLS)

    ws.column_dimensions["A"].width = 42.0
    for c in range(2, last_val_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 15.0

    # ── title ─────────────────────────────────────────────────────────────────
    paint_row(ws, 1, 1, last_col, font=F_TITLE, fill=GOLD_PALE, border=B_BAND,
              align=AL_LEFT)
    style_cell(ws, 1, 1, "WEEKLY REPORT AS ON ")
    style_cell(ws, 1, 2, as_of, fmt=r'[$-409]mmmm\ d\,\ yyyy')
    ws.row_dimensions[1].height = 30

    # ── market statistics ─────────────────────────────────────────────────────
    row = 3
    paint_row(ws, row, 1, last_col, font=F_HDR_DARK, fill=AMBER, border=B_HDR,
              align=AL_CENTER_WRAP)
    style_cell(ws, row, 1, "Market Statistics", align=AL_LEFT)
    style_cell(ws, row, 2, "Current")
    style_cell(ws, row, 4, "Previous Week")
    style_cell(ws, row, 6, mar31, fmt=r'd" "mmm" "yyyy')
    style_cell(ws, row, 7, "WEEK %")
    style_cell(ws, row, 8, "YTD %")
    row += 1

    by_code = {b["code"]: b for b in (benchmarks or [])}
    for code, label in BENCHMARK_ROWS:
        b = by_code.get(code) or {}
        paint_row(ws, row, 1, last_col, font=F_BODY, fill=NO_FILL, border=B_THIN)
        style_cell(ws, row, 1, label, font=F_BODY_BOLD, align=AL_LEFT)
        style_cell(ws, row, 2, b.get("current"),   fmt=PRICE, align=AL_RIGHT)
        style_cell(ws, row, 4, b.get("prev_week"), fmt=PRICE, align=AL_RIGHT)
        style_cell(ws, row, 6, b.get("mar31"),     fmt=PRICE, align=AL_RIGHT)
        style_cell(ws, row, 7, b.get("week_pct"),  fmt=PCT2,  align=AL_RIGHT)
        style_cell(ws, row, 8, b.get("ytd_pct"),   fmt=PCT2,  align=AL_RIGHT)
        row += 1

    # ── asset monitor ─────────────────────────────────────────────────────────
    row += 1
    paint_row(ws, row, 1, last_col, font=F_TITLE16, fill=NO_FILL, border=B_THIN,
              align=AL_LEFT)
    style_cell(ws, row, 1, "ASSET MONITOR FOR COMBINED HDR GROUP")
    style_cell(ws, row, last_val_col, "Rs in Actual", font=F_NOTE, align=AL_RIGHT)
    row += 1

    # group banner row
    paint_row(ws, row, 1, last_col, font=F_HDR12, fill=GOLD_DARK, border=B_BAND,
              align=AL_CENTER)
    style_cell(ws, row, 1, "Asset Class", align=AL_LEFT)
    for i, blk in enumerate(blocks):
        c0 = 2 + i * 5
        ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c0 + 4)
        ws.cell(row=row, column=c0).value = blk["label"]
    ws.row_dimensions[row].height = 22
    row += 1

    # per-group column headers (two rows: the YTD qualifier sits on the second)
    hdr, hdr2 = row, row + 1
    for rr in (hdr, hdr2):
        paint_row(ws, rr, 1, last_col, font=F_HDR, fill=GOLD_DARK, border=B_HDR,
                  align=AL_CENTER_WRAP)
    ws.merge_cells(start_row=hdr, start_column=1, end_row=hdr2, end_column=1)
    for i, blk in enumerate(blocks):
        c0 = 2 + i * 5
        for j, name in enumerate(AE_GROUP_COLS):
            c = c0 + j
            if name in ("Profit & Loss", "Returns (%)"):
                style_cell(ws, hdr,  c, name, font=F_HDR)
                style_cell(ws, hdr2, c, "YTD", font=F_HDR)
            else:
                ws.merge_cells(start_row=hdr, start_column=c, end_row=hdr2, end_column=c)
                style_cell(ws, hdr, c,
                           "Total CMV" if (name == "CMV" and i == n_blocks - 1) else name,
                           font=F_HDR)
    for j, (name, hidden) in enumerate(AE_TAIL_COLS):
        c = last_val_col + 1 + j
        ws.merge_cells(start_row=hdr, start_column=c, end_row=hdr2, end_column=c)
        style_cell(ws, hdr, c, name, font=F_HDR)
        ws.column_dimensions[get_column_letter(c)].width = 13.0
        ws.column_dimensions[get_column_letter(c)].hidden = hidden
    ws.row_dimensions[hdr].height = 28
    row = hdr2 + 1

    note(ws, row, last_val_col + len(AE_TAIL_COLS),
         "all YTD returns in this sheet are weighted averages of the individual "
         "portfolio %")
    row += 1

    # ── the lettered lines ────────────────────────────────────────────────────
    for label, spec in AE_LINES:
        if label is None:
            row += 1
            continue
        is_total = label in _AE_TOTAL_LINES
        font = F_GRANDTOT if is_total else F_BODY_BOLD
        fill = GOLD_DARK  if is_total else NO_FILL
        paint_row(ws, row, 1, last_col, font=font, fill=fill,
                  border=B_BAND if is_total else B_THIN)
        style_cell(ws, row, 1, label, font=font, align=AL_LEFT)

        for i, blk in enumerate(blocks):
            c0 = 2 + i * 5
            if isinstance(spec, str) and spec.startswith("realised:"):
                # Realised P&L is a P&L-only line — it has no market value.
                val = (blk.get("realised") or {}).get(spec.split(":", 1)[1])
                style_cell(ws, row, c0 + 2, val or None, fmt=MONEY, align=AL_RIGHT)
                continue
            rows = _rows_of(blk["sec"], spec)
            cmv  = _sum(rows, "current_value")
            cost = _sum(rows, "cost")
            pnl  = _sum(rows, "pnl_ytd")
            wk, prev = _weekly_agg(rows)
            fmt  = MONEY_TOT if is_total else MONEY
            style_cell(ws, row, c0,     cmv,  fmt=fmt,  align=AL_RIGHT)
            style_cell(ws, row, c0 + 1, wk,   fmt=fmt,  align=AL_RIGHT)
            style_cell(ws, row, c0 + 2, pnl,  fmt=fmt,  align=AL_RIGHT)
            style_cell(ws, row, c0 + 3,
                       (wk / prev) if (wk is not None and prev) else None,
                       fmt=PCT2, align=AL_RIGHT)
            style_cell(ws, row, c0 + 4,
                       (pnl / (cmv - pnl)) if (pnl is not None and cmv and (cmv - pnl))
                       else None,
                       fmt=PCT2, align=AL_RIGHT)
            if is_total:
                style_cell(ws, row, last_val_col + 1, cost, fmt=MONEY_TOT,
                           align=AL_RIGHT)
        row += 1

    row += 1
    row = note(ws, row, 1, "*YTD P&L and Returns do not include Non-Earning Assets",
               span=last_val_col)
    row += 1
    row = note(ws, row, 1, "REASONING BEHIND THE SHEET", span=last_val_col)
    row = note(ws, row, 1, "Course correction for asset class as a whole",
               span=last_val_col)
    note(ws, row, 1,
         "*Realised p&L figures are shown separetely but they have been already "
         "included under column D for each asset class because it is picked from the "
         "individual sheets which includes that figure in the totals for each asset "
         "class ", span=last_val_col)
    ws.freeze_panes = ws.cell(row=hdr2 + 1, column=2)
    return ws


# ══════════════════════════════════════════════════════════════════════════════
#  FY Realised Profit & Loss  —  the client's "FY2627" page, cloned per entity
# ══════════════════════════════════════════════════════════════════════════════

RP_LAST = 7
RP_WIDTHS = {"A": 47.71, "B": 11.57, "C": 15.00, "D": 11.57, "E": 15.00,
             "F": 15.00, "G": 10.00}


def _realised_buckets(realised: list) -> dict:
    """Split realised rows into the client's own report sections."""
    out = defaultdict(list)
    for r in realised or []:
        g   = (r.get("group") or "").strip()
        cat = (r.get("category") or "").strip()
        name = r.get("security_name") or ""
        # Arbitrage / equity-savings funds are tagged MF_EQUITY upstream but the
        # client reports them under Fixed Income, so the name is checked before the
        # group — the same precedence classify_bundle() uses for the weekly page.
        if _has(name, _ARBITRAGE_HINTS):
            out["fi_arbitrage"].append(r)
        elif g == "Fixed Income":
            out[f"fi_{fixed_income_bucket({'security_name': name})}"].append(r)
        elif g == "PMS" or cat == "PMS":
            out["pms"].append(r)
        elif g == "Foreign Equity":
            out["foreign"].append(r)
        elif g == "Commodities":
            out["gold"].append(r)
        elif g == "Alternates":
            out["unlisted"].append(r)
        elif cat == "Mutual Funds":
            out["mf_equity"].append(r)
        else:
            out["direct"].append(r)
    return out


class _RealisedWriter:
    def __init__(self, ws):
        self.ws = ws
        self.row = 1

    def blank(self, n=1):
        self.row += n

    def title(self, text, *, font=F_TITLE, fill=NO_FILL, height=28, span=RP_LAST):
        r = self.row
        paint_row(self.ws, r, 1, span, font=font, fill=fill, align=AL_CENTER)
        self.ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
        self.ws.cell(row=r, column=1).value = text
        self.ws.row_dimensions[r].height = height
        self.row += 1

    def headers(self, cols):
        r = self.row
        paint_row(self.ws, r, 1, RP_LAST, font=F_HDR, fill=GOLD_DARK, border=B_HDR,
                  align=AL_CENTER_WRAP)
        for c, text in enumerate(cols, 1):
            self.ws.cell(row=r, column=c).value = text
        self.ws.row_dimensions[r].height = 26
        self.row += 1

    def _row(self, vals, *, font, fill, border, fmt=MONEY, date_cols=(), pct_cols=()):
        r = self.row
        paint_row(self.ws, r, 1, RP_LAST, font=font, fill=fill, border=border)
        for c, v in enumerate(vals, 1):
            cell = style_cell(self.ws, r, c, v)
            if c == 1:
                cell.alignment = AL_LEFT_WRAP
            elif c in date_cols:
                cell.number_format = DATE_DMY; cell.alignment = AL_CENTER
            elif c in pct_cols:
                cell.number_format = PCT2;     cell.alignment = AL_RIGHT
            else:
                cell.number_format = fmt;      cell.alignment = AL_RIGHT
        self.row += 1
        return r

    def summary_row(self, label, rows, *, total=False):
        pa  = _sum(rows, "purchase_amount")
        sa  = _sum(rows, "sale_amount")
        pnl = _sum(rows, "pnl")
        ret = (pnl / pa) if (pnl is not None and pa) else None
        return self._row([label, pa, sa, pnl, ret if ret is not None else "-",
                          None, None],
                         font=F_GRANDTOT if total else F_BODY_BOLD,
                         fill=GOLD_DARK if total else NO_FILL,
                         border=B_BAND if total else B_THIN,
                         fmt=MONEY_TOT if total else MONEY, pct_cols=(5,))

    def detail_total(self, label, rows, *, total=False):
        """Section total, written ABOVE its detail rows the way the client does."""
        pa  = _sum(rows, "purchase_amount")
        sa  = _sum(rows, "sale_amount")
        pnl = _sum(rows, "pnl")
        ret = (pnl / pa) if (pnl is not None and pa) else None
        return self._row([label, None, pa, None, sa, pnl,
                          ret if ret is not None else "-"],
                         font=F_GRANDTOT if total else F_SUBTOTAL,
                         fill=GOLD_DARK if total else GOLD_LIGHT,
                         border=B_BAND if total else B_THIN,
                         fmt=MONEY_TOT, date_cols=(2, 4), pct_cols=(7,))

    def detail(self, r):
        pnl = _f(r, "pnl")
        pa  = _f(r, "purchase_amount")
        ret = _f(r, "return_pct")
        if ret is None and pnl is not None and pa:
            ret = pnl / pa
        return self._row([r.get("security_name"), r.get("purchase_date"), pa,
                          r.get("sale_date"), _f(r, "sale_amount"), pnl,
                          ret if ret is not None else "-"],
                         font=F_BODY, fill=NO_FILL, border=B_THIN,
                         date_cols=(2, 4), pct_cols=(7,))

    def caption(self, text):
        r = self.row
        paint_row(self.ws, r, 1, RP_LAST, font=F_BODY, fill=NO_FILL, border=B_THIN)
        style_cell(self.ws, r, 1, text, font=F_BODY_BOLD, align=AL_LEFT)
        self.row += 1

    def section(self, label, rows, *, total=False):
        self.detail_total(label, rows, total=total)
        for r in rows:
            self.detail(r)


def build_realised_pnl(ws, label: str, as_of: date, realised: list = None,
                       dividends: list = None):
    """
    The client's FY realised-gains page for one entity or group:
      summary → '- FIXED INCOME INVESTMENT' → '- EQUITIES' → '- ALTERNATES'
      → 'DIVIDEND RECIEVED'
    """
    ws.sheet_view.showGridLines = False
    set_widths(ws, RP_WIDTHS)
    b  = _realised_buckets(realised)
    fy = _fy_label(as_of)
    w  = _RealisedWriter(ws)

    DETAIL_COLS = ["ASSET CLASS", "Purchase Date", "Purchase Amount", "Sale Date",
                   "Sale Amount", "Profit/Loss", "Returns"]

    fixed_rows  = _rows_of(b, ["fi_liquid", "fi_debt", "fi_arbitrage", "fi_ppf"])
    equity_rows = _rows_of(b, ["mf_equity", "pms", "direct", "foreign"])
    alt_rows    = _rows_of(b, ["unlisted", "gold"])

    # ── summary ───────────────────────────────────────────────────────────────
    w.title(f"ALL Income Monitor - {label.upper()}", fill=GOLD_PALE)
    w.title(f"Realised Invesment Gains or Losses for FY {fy}",
            font=F_SECTION, fill=CREAM, height=22)
    w.headers(["ASSET CLASS", "Purchase Amount", "Sale Amount", "Profit/Loss",
               "Returns", None, None])
    w.summary_row("Long Term Equity (Including dividend recieved)", equity_rows)
    w.summary_row("Fixed Income", fixed_rows)
    w.summary_row("SUB TOTAL (A)", equity_rows + fixed_rows, total=True)
    w.blank()
    w.summary_row("ALTERNATES (B)", alt_rows, total=True)
    w.summary_row("Unlisted", b["unlisted"])
    w.summary_row("Gold ETF", b["gold"])
    w.blank()
    w.summary_row("GRAND TOTAL (A+B)", equity_rows + fixed_rows + alt_rows, total=True)
    w.blank(2)

    # ── fixed income detail ───────────────────────────────────────────────────
    w.title("- FIXED INCOME INVESTMENT")
    w.title(label.upper(), font=F_BANNER, fill=GOLD_DARK, height=24)
    w.headers(DETAIL_COLS)
    w.section("MF-LIQUID FUND TOTAL", b["fi_liquid"])
    w.blank()
    w.section("DEBT FUND TOTAL",      b["fi_debt"])
    w.blank()
    w.section("ARBITRAGE FUND TOTAL", b["fi_arbitrage"])
    w.blank()
    w.section("PUBLIC PROVIDENT FUND", b["fi_ppf"])
    w.blank()
    w.detail_total("A. TOTAL FIXED INVESMENT", fixed_rows, total=True)
    w.blank(2)

    # ── equities detail ───────────────────────────────────────────────────────
    w.title("- EQUITIES")
    w.title(label.upper(), font=F_BANNER, fill=GOLD_DARK, height=24)
    w.headers(DETAIL_COLS)
    w.section("MF EQUITY TOTAL (i)",   b["mf_equity"])
    w.blank()
    w.section("PMS EQUITY TOTAL (ii)", b["pms"])
    w.blank()
    # The client lists direct-equity realisations under the demat they happened in.
    w.detail_total("DIRECT EQUITIES DESK (iii)", b["direct"])
    by_broker = defaultdict(list)
    for r in b["direct"]:
        # Broker keys arrive as feed slugs ('angel_one'); the client writes them as
        # plain caps ('ANGEL ONE').
        by_broker[(r.get("broker") or "").replace("_", " ").upper() or "OTHER"].append(r)
    for broker in sorted(by_broker):
        w.caption(broker)
        for r in by_broker[broker]:
            w.detail(r)
    w.blank()
    w.section("FOREIGN EQUITY FUNDS (iv)", b["foreign"])
    w.blank()
    w.detail_total("B. TOTAL EQUITY EXPOSURE (i+ii+iii+iv)", equity_rows, total=True)
    w.blank()
    w.detail_total("C. SUB GRAND TOTAL(A+B)", equity_rows + fixed_rows, total=True)
    w.blank(2)

    # ── alternates detail ─────────────────────────────────────────────────────
    w.title("- ALTERNATES")
    w.title(label.upper(), font=F_BANNER, fill=GOLD_DARK, height=24)
    w.headers(DETAIL_COLS)
    w.section("Unlisted Shares", b["unlisted"])
    w.blank()
    w.section("GOLD ETF TOTAL",  b["gold"])
    w.blank()
    w.detail_total("D. TOTAL ALTERNATES", alt_rows, total=True)
    w.blank(2)

    # ── dividends ─────────────────────────────────────────────────────────────
    w.title("DIVIDEND RECIEVED ")
    w.title(label.upper(), font=F_BANNER, fill=GOLD_DARK, height=24)
    w.headers(["SCRIPT NAME", "Dividend Date", "Dividend Amount", None, None,
               None, None])
    div_by_broker = defaultdict(list)
    for d in (dividends or []):
        div_by_broker[(d.get("broker") or "").replace("_", " ").upper()
                      or "OTHER"].append(d)
    for broker in sorted(div_by_broker):
        w.caption(broker)
        for d in div_by_broker[broker]:
            w._row([d.get("security_name"), d.get("pay_date") or d.get("date"),
                    _f(d, "amount"), None, None, None, None],
                   font=F_BODY, fill=NO_FILL, border=B_THIN, date_cols=(2,))
    w.detail_total("C. SUB GRAND TOTAL", [], total=True)
    r = w.row - 1
    style_cell(ws, r, 3, _sum(dividends or [], "amount"), fmt=MONEY_TOT,
               align=AL_RIGHT)
    return ws
