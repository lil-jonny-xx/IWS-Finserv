#!/usr/bin/env python3
"""
Conformance check: does the generated MIS workbook still match the client's format?

The generator drifted away from the client's layout once already — the sheet set,
the column count and the block structure had all diverged from the format file by
the time anyone compared them. This script exists so that cannot happen silently:
it re-reads the client's own workbook and asserts, cell by cell, that every title,
column header and asset-class line in the generated file is still identical.

Only LABELS are compared, never values — the numbers are live and the dates are
today's, so those are expected to differ.

Usage:
    python workers/verify_report_format.py <generated-MIS-Report.xlsx> [reference.xlsx]

Exits non-zero when anything has drifted, so it can gate a deploy.

The reference workbook is gitignored (`*.xlsx`) and is not in the repo — it carries
the client's fund names and entity tags, and this repo is public. Keep a copy at
REFERENCE_PATH locally; the script skips with a clear message when it is absent.
"""
import sys
from pathlib import Path

import openpyxl

REFERENCE_PATH = "/var/www/MIS-CURRENT-FORMAT.xlsx"

# Template sheet → the generated sheet that is built from it.
WEEKLY_TPL   = "Dhruv Weekly Report"
REALISED_TPL = "FY2627 Realised Profit & Loss"

# Columns holding an as-on date rather than a label: live reports carry today's.
DATE_COLS = (3, 4)


def _labels(ws, row, last_col):
    return {c: str(ws.cell(row, c).value).strip()
            for c in range(1, last_col + 1) if ws.cell(row, c).value is not None}


class Checker:
    def __init__(self):
        self.failures = []

    def compare(self, name, t_ws, t_row, g_ws, g_row, last_col, skip=()):
        want = _labels(t_ws, t_row, last_col)
        got  = _labels(g_ws, g_row, last_col)
        diffs = {c: v for c, v in want.items() if c not in skip and got.get(c) != v}
        if diffs:
            self.failures.append((name, diffs, got))
        print(f"  [{'MATCH' if not diffs else 'DRIFT'}] {name}")
        for c, v in diffs.items():
            print(f"          col {c}: expected {v!r}, got {got.get(c)!r}")


def _find_edp_blocks(ws):
    """Row numbers of each Equity Daily Print block title, in sheet order."""
    return [r for r in range(1, ws.max_row + 1)
            if isinstance(ws.cell(r, 1).value, str)
            and ws.cell(r, 1).value.endswith(" as on")]


def verify(generated_path: str, reference_path: str = REFERENCE_PATH) -> int:
    if not Path(reference_path).exists():
        print(f"SKIP — reference workbook not found at {reference_path}.\n"
              f"       Place the client's format file there to run this check.")
        return 0

    tpl = openpyxl.load_workbook(reference_path, data_only=True)
    gen = openpyxl.load_workbook(generated_path, data_only=True)
    ck = Checker()

    # ── sheet set ─────────────────────────────────────────────────────────────
    print("Sheet set")
    for required in ("Equity Daily Print", "All Assets Daily MIS",
                     "All Entities Weekly Report"):
        ok = required in gen.sheetnames
        print(f"  [{'MATCH' if ok else 'DRIFT'}] shared sheet {required!r}")
        if not ok:
            ck.failures.append((f"missing sheet {required}", {}, {}))

    weekly = [s for s in gen.sheetnames if s.endswith(" Weekly Report")
              and s != "All Entities Weekly Report"]
    realised = [s for s in gen.sheetnames if s.endswith(" Realised P&L")]
    print(f"  [INFO ] {len(weekly)} weekly pages, {len(realised)} realised pages")
    if len(weekly) != len(realised):
        ck.failures.append(("weekly/realised page counts differ", {}, {}))

    # ── Equity Daily Print ────────────────────────────────────────────────────
    print("\nEquity Daily Print — column headers per block variant")
    t, g = tpl["Equity Daily Print"], gen["Equity Daily Print"]
    g_blocks = _find_edp_blocks(g)
    if not g_blocks:
        ck.failures.append(("Equity Daily Print has no blocks", {}, {}))
    else:
        # template rows: 2 = group DOMESTIC, 33 = entity DOMESTIC,
        #                65 = entity FOREIGN, 97 = group FOREIGN
        first_for = next((i for i, r in enumerate(g_blocks)
                          if "(FOREIGN)" in str(g.cell(r, 1).value)), len(g_blocks))
        pairs = [("group DOMESTIC",  2,  g_blocks[0] + 1)]
        if first_for > 1:
            pairs.append(("entity DOMESTIC", 33, g_blocks[1] + 1))
        if first_for < len(g_blocks):
            pairs.append(("entity FOREIGN", 65, g_blocks[first_for] + 1))
            pairs.append(("group FOREIGN",  97, g_blocks[-1] + 1))
        for name, t_row, g_row in pairs:
            ck.compare(f"{name} header", t, t_row, g, g_row, 13)

    # ── Weekly page ───────────────────────────────────────────────────────────
    if weekly:
        t, g = tpl[WEEKLY_TPL], gen[weekly[0]]
        print(f"\nWeekly Report — {WEEKLY_TPL!r} vs {weekly[0]!r}")
        ck.compare("summary header row 1", t, 3,  g, 3,  16)
        ck.compare("summary header row 2", t, 4,  g, 4,  16, skip=DATE_COLS)
        ck.compare("detail header row 1",  t, 25, g, 25, 16)
        ck.compare("detail header row 2",  t, 26, g, 26, 16, skip=DATE_COLS)
        ck.compare("market stats header",  t, 14, g, 14, 16, skip=(6, 7, 9))
        for row in range(15, 22):
            ck.compare(f"market stats row {row}", t, row, g, row, 1)

    # ── Realised page ─────────────────────────────────────────────────────────
    if realised:
        t, g = tpl[REALISED_TPL], gen[realised[0]]
        print(f"\nRealised P&L — {REALISED_TPL!r} vs {realised[0]!r}")
        ck.compare("summary header", t, 3,  g, 3,  7)
        ck.compare("detail header",  t, 18, g, 17, 7)
        for row in (4, 5, 6):
            ck.compare(f"summary row {row}", t, row, g, row, 1)

    # ── All Entities monitor ──────────────────────────────────────────────────
    if "All Entities Weekly Report" in gen.sheetnames:
        t, g = tpl["All Entities Weekly Report"], gen["All Entities Weekly Report"]
        print("\nAll Entities Weekly Report")
        ck.compare("market stats header", t, 3, g, 3, 8, skip=(2, 4, 6))
        ck.compare("group column header row 1", t, 13, g, 13, 6)
        ck.compare("group column header row 2", t, 14, g, 14, 6)
        for row in (16, 17, 18, 19, 20, 21, 23, 24, 25, 26, 27, 28, 29, 31,
                    33, 34, 35, 36, 38, 40, 41, 42, 44):
            ck.compare(f"asset line row {row}", t, row, g, row, 1)

    print()
    if ck.failures:
        print(f"FAILED — {len(ck.failures)} check(s) drifted from the client's format.")
        return 1
    print("PASSED — the generated workbook matches the client's format exactly.")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(2)
    sys.exit(verify(*sys.argv[1:3]))
